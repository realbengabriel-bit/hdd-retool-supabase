# -*- coding: utf-8 -*-
r"""
OIF/EH PDF generator front-layer v3.7 employment 9.6 + guest worker 9.7 annex support

Purpose
-------
This layer runs BEFORE the existing EnterHungary assistant.
It converts a Retool/Supabase OIF/EH package payload into:
  1) an extractor-compatible, filled source PDF for the existing eh_pdf_extractor_v8 script,
  2) a TSV row compatible with ENTERHUNGARY_20260519_fin.xlsm / Munkavállalók,
  3) optional XLSM copy with the generated row inserted and Start!C1 set,
  4) optional QR back page,
  5) optional Supabase Storage upload + generated_files DB registration.

This does NOT replace the existing EH web automation.
The existing EH automation still handles:
  - EH form filling from XLSM
  - attachment upload from person/project folder

Supported existing EH automation types
--------------------------------------
The supplied legacy scripts support:
  - vendeg: 9.7 Vendégmunkás betétlap
  - nemzeti: 9.12 Nemzeti Kártya betétlap

9.6 Foglalkoztatási célú is detected, but returned as unsupported_for_legacy_eh_agent
until a new mapping / EH assistant type is added.

FastAPI usage
-------------
  uvicorn eh_oif_pdf_generator_frontlayer_v1:app --host 127.0.0.1 --port 8787

POST /generate-oif-package-pdfs
{
  "package_id": "...",
  "python_export_row": {...},
  "person_folder": "C:/.../NAME",
  "workbook_path": "C:/.../ENTERHUNGARY_20260519_fin.xlsm",
  "upload_generated_to_storage": true
}

CLI usage
---------
  py eh_oif_pdf_generator_frontlayer_v1.py --payload-json payload.json --out-dir runs/test --workbook ENTERHUNGARY_20260519_fin.xlsm

Environment for Supabase upload
-------------------------------
  SUPABASE_URL=https://xxxxx.supabase.co
  SUPABASE_SERVICE_ROLE_KEY=...
  OIF_EH_GENERATED_BUCKET=oif-eh-generated

Dependencies
------------
  py -m pip install fastapi uvicorn reportlab qrcode[pil] pymupdf openpyxl requests
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import re
import shutil
import sys
import tempfile
import traceback
import uuid
from dataclasses import dataclass, asdict
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
except Exception:  # pragma: no cover
    print("Missing dependency: reportlab. Install: py -m pip install reportlab", file=sys.stderr)
    raise

PDF_FONT = "Helvetica"
PDF_FONT_BOLD = "Helvetica-Bold"

def _register_unicode_font() -> None:
    global PDF_FONT, PDF_FONT_BOLD
    candidates = [
        os.environ.get("OIF_EH_PDF_FONT", ""),
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        "/usr/share/fonts/truetype/noto/NotoSans-Regular.ttf",
        "C:/Windows/Fonts/arial.ttf",
        "C:/Windows/Fonts/Arial.ttf",
    ]
    regular = next((x for x in candidates if x and Path(x).exists() and "Bold" not in x and "bold" not in x), "")
    bold_candidates = [
        os.environ.get("OIF_EH_PDF_FONT_BOLD", ""),
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        "/usr/share/fonts/truetype/noto/NotoSans-Bold.ttf",
        "C:/Windows/Fonts/arialbd.ttf",
        "C:/Windows/Fonts/Arialbd.ttf",
    ]
    bold = next((x for x in bold_candidates if x and Path(x).exists()), "")
    try:
        if regular:
            pdfmetrics.registerFont(TTFont("OIFUnicode", regular))
            PDF_FONT = "OIFUnicode"
        if bold:
            pdfmetrics.registerFont(TTFont("OIFUnicode-Bold", bold))
            PDF_FONT_BOLD = "OIFUnicode-Bold"
        elif regular:
            PDF_FONT_BOLD = PDF_FONT
    except Exception:
        PDF_FONT = "Helvetica"
        PDF_FONT_BOLD = "Helvetica-Bold"

_register_unicode_font()

try:
    import qrcode
except Exception:  # pragma: no cover
    qrcode = None

try:
    import fitz  # PyMuPDF, only used for optional verification/extractor fallback
except Exception:  # pragma: no cover
    fitz = None

try:
    import openpyxl
except Exception:  # pragma: no cover
    openpyxl = None

try:
    import requests
except Exception:  # pragma: no cover
    requests = None

try:
    from fastapi import FastAPI
    from pydantic import BaseModel, Field
except Exception:  # pragma: no cover
    FastAPI = None
    BaseModel = object
    Field = lambda default=None, **kwargs: default  # type: ignore

# -----------------------------
# Column list compatible with the user's v8 extractor / XLSM row
# -----------------------------

def column_letter(col_num: int) -> str:
    s = ""
    n = col_num
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def make_letters(last: str = "DH") -> List[str]:
    out = []
    n = 1
    while True:
        col = column_letter(n)
        out.append(col)
        if col == last:
            return out
        n += 1

LETTERS = make_letters("DH")

# -----------------------------
# Utilities
# -----------------------------

def now_id() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S") + "_" + uuid.uuid4().hex[:8]


def one_line(value: Any) -> str:
    if value is None:
        return ""
    s = str(value).replace("\u00a0", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def norm(value: Any) -> str:
    s = one_line(value).lower()
    table = str.maketrans("áéíóöőúüűÁÉÍÓÖŐÚÜŰ", "aeiooouuuAEIOOOUUU")
    return s.translate(table)


def parse_date_to_dot(value: Any, fallback: str = "") -> str:
    s = one_line(value)
    if not s:
        return fallback
    m = re.search(r"(\d{4})[.\-/ ]+(\d{1,2})[.\-/ ]+(\d{1,2})", s)
    if m:
        return f"{int(m.group(1)):04d}.{int(m.group(2)):02d}.{int(m.group(3)):02d}"
    return fallback


def parse_date_to_iso(value: Any, fallback: str = "") -> str:
    d = parse_date_to_dot(value, fallback="")
    if d:
        return d.replace(".", "-")
    return fallback


def money_digits(value: Any, fallback: str = "455976") -> str:
    s = one_line(value)
    m = re.search(r"([\d\s.]+)", s)
    if not m:
        return fallback
    out = re.sub(r"\D", "", m.group(1))
    return out or fallback


def slugify(value: Any, max_len: int = 80) -> str:
    s = norm(value)
    s = re.sub(r"[^a-z0-9]+", "_", s).strip("_")
    return (s[:max_len] or "oif_eh")


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def flatten_json(obj: Any, prefix: str = "") -> Dict[str, Any]:
    out: Dict[str, Any] = {}
    if isinstance(obj, dict):
        for k, v in obj.items():
            key = f"{prefix}.{k}" if prefix else str(k)
            out[key] = v
            out.update(flatten_json(v, key))
    elif isinstance(obj, list):
        for i, v in enumerate(obj[:20]):
            key = f"{prefix}[{i}]"
            out[key] = v
            out.update(flatten_json(v, key))
    return out


def parse_maybe_json(value: Any) -> Any:
    if isinstance(value, (dict, list)):
        return value
    if not isinstance(value, str):
        return value
    s = value.strip()
    if not s:
        return {}
    try:
        return json.loads(s)
    except Exception:
        return value


def first_value(*values: Any, default: str = "") -> str:
    for v in values:
        if isinstance(v, list) and v:
            v = v[0]
        if isinstance(v, dict):
            continue
        s = one_line(v)
        if s:
            return s
    return default


def pick(flat: Dict[str, Any], *keys: str, default: str = "") -> str:
    # exact keys first
    for k in keys:
        if k in flat and one_line(flat[k]):
            return one_line(flat[k])
    # suffix / last segment match
    wanted = [norm(k).replace(".", "_") for k in keys]
    for k, v in flat.items():
        nk = norm(k).replace(".", "_")
        if any(nk.endswith(w) or w.endswith(nk) for w in wanted) and one_line(v):
            return one_line(v)
    return default


def split_full_name(full_name: str) -> Tuple[str, str]:
    parts = [p for p in one_line(full_name).split(" ") if p]
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], " ".join(parts[1:])


def infer_type(process_kind: str, package_type: str, payload: Dict[str, Any]) -> str:
    hay = norm(" ".join([process_kind, package_type, json.dumps(payload, ensure_ascii=False)[:5000]]))
    if any(x in hay for x in ["9.6", "9.6", "foglalkoztatasi celu", "employment purpose", "foglalkoztatas"]):
        return "foglalkoztatasi"  # legacy scripts do not support this yet
    if any(x in hay for x in ["nemzeti", "national_card", "national card", "9.12", "kartya"]):
        return "nemzeti"
    if any(x in hay for x in ["vendeg", "guest", "9.7", "extension", "hosszabbit", "feor"]):
        return "vendeg"
    return "vendeg"


def normalize_nationality(value: str) -> str:
    n = norm(value)
    if "fulop" in n or "philipp" in n:
        return "Fülöp Szigeteki"
    if "ukran" in n or "ukraj" in n:
        return "Ukrán"
    if "kazah" in n or "kazak" in n:
        return "Kazah"
    if "orosz" in n or "russian" in n:
        return "Orosz"
    if "mold" in n:
        return "Moldáv"
    return value or ""


def normalize_country_from_nationality(value: str) -> str:
    n = norm(value)
    if "fulop" in n or "philipp" in n:
        return "Fülöp-Szigetek"
    if "ukran" in n or "ukraj" in n:
        return "Ukrajna"
    if "kazah" in n or "kazak" in n:
        return "Kazahsztán"
    if "orosz" in n:
        return "Oroszország"
    if "mold" in n:
        return "Moldova"
    return value or ""


def ensure_dir(path: Path) -> Path:
    path.mkdir(parents=True, exist_ok=True)
    return path

# -----------------------------
# Data model
# -----------------------------

@dataclass
class GeneratedPerson:
    app_type: str
    full_name: str
    family_name: str
    given_name: str
    birth_family_name: str
    birth_given_name: str
    gender: str
    birth_place: str
    birth_country: str
    birth_date_dot: str
    nationality: str
    mother_family: str
    mother_given: str
    passport_no: str
    passport_issue_place: str
    passport_issue_date_dot: str
    passport_valid_until_dot: str
    foreign_country: str
    foreign_city: str
    foreign_street: str
    hu_zip: str
    hu_city: str
    hu_street: str
    hu_stype: str
    hu_house: str
    application_until_dot: str
    kelt_dot: str
    agreement_date_dot: str
    feor_code: str
    feor_name: str
    employer_name: str
    company_code: str
    employer_short: str
    tax_no: str
    ksh: str
    teaor: str
    salary: str
    work_addr: str
    language: str
    reg_date_words: str
    reg_no: str
    eh_case_id: str
    package_id: str
    workflow_case_id: str
    contact_phone: str
    contact_email: str
    employer_zip: str
    employer_city: str
    employer_street: str
    employer_stype: str
    employer_house: str
    employer_building: str
    employer_staircase: str
    employer_floor: str
    employer_door: str
    source_payload: Dict[str, Any]


def person_from_payload(input_payload: Dict[str, Any]) -> GeneratedPerson:
    row = input_payload.get("python_export_row") or input_payload.get("export_row") or input_payload
    if isinstance(row, list):
        row = row[0] if row else {}
    row = dict(row or {})
    parsed_payload = parse_maybe_json(row.get("python_payload_json") or row.get("python_payload") or row.get("payload") or {})
    payload = parsed_payload if isinstance(parsed_payload, dict) else {}
    merged = {**row, "payload": payload}
    flat = flatten_json(merged)

    process_kind = first_value(input_payload.get("process_kind"), row.get("process_kind"), pick(flat, "process_kind"))
    package_type = first_value(input_payload.get("package_type"), row.get("package_type"), pick(flat, "package_type"))
    app_type = first_value(input_payload.get("app_type"), input_payload.get("application_type"), default="")
    if app_type not in {"vendeg", "nemzeti", "foglalkoztatasi"}:
        app_type = infer_type(process_kind, package_type, merged)

    full_name = first_value(
        input_payload.get("full_name"),
        row.get("primary_person_name"), row.get("full_name"), row.get("person_name"),
        pick(flat, "primary_person_name", "full_name", "person.name", "name"),
        default="ISMERETLEN SZEMÉLY",
    )
    family = first_value(row.get("last_name"), pick(flat, "last_name", "family_name"), default="")
    given = first_value(row.get("first_name"), pick(flat, "first_name", "given_name"), default="")
    if not family or not given:
        f, g = split_full_name(full_name)
        family = family or f
        given = given or g

    nationality = normalize_nationality(first_value(row.get("nationality"), pick(flat, "nationality", "citizenship"), default=""))
    birth_country = normalize_country_from_nationality(first_value(pick(flat, "birth_country", "country_of_birth"), nationality, default=""))

    employer_name = first_value(row.get("employer_name"), pick(flat, "employer_name", "employer", "company_name"), default="HD DIREKT HUNGARY KFT.")
    employer_upper = employer_name.upper()
    if "HR DIREKT" in employer_upper:
        company_code = "HRD"
    elif "HD DIREKT" in employer_upper or employer_upper.startswith("HDD"):
        company_code = "HDD"
    else:
        company_code = ""
    employer_short = "HR Direkt Kft." if company_code == "HRD" else ("HD Direkt Hungary Kft." if company_code == "HDD" else employer_name)

    # FEOR-mód rákötés: Retool az új UAHUN workflow FEOR módosításából
    # top-level és python_export_row / python_payload szinten is küldhet effective_feor_* mezőket.
    # Ezek erősebbek, mint a csomag eredeti FEOR-ja.
    feor_code = re.sub(r"\.0$", "", first_value(
        input_payload.get("effective_feor_code"),
        input_payload.get("oif_eh_effective_feor_code"),
        input_payload.get("feor_code"),
        row.get("effective_feor_code"),
        row.get("oif_eh_effective_feor_code"),
        row.get("feor_code"),
        row.get("feor"),
        pick(flat, "effective_feor_code", "oif_eh_effective_feor_code", "feor_code", "feor"),
        default="9310"
    ))
    feor_name = first_value(
        input_payload.get("effective_feor_name"),
        input_payload.get("oif_eh_effective_feor_name"),
        input_payload.get("feor_name"),
        row.get("effective_feor_name"),
        row.get("oif_eh_effective_feor_name"),
        row.get("feor_name"),
        row.get("position_name"),
        row.get("job_title"),
        pick(flat, "effective_feor_name", "oif_eh_effective_feor_name", "feor_name", "position_name", "job_title"),
        default="Egyszerű ipari foglalkozású" if feor_code == "9310" else "Betanított munkakör"
    )

    kelt_dot = parse_date_to_dot(first_value(input_payload.get("generated_at"), row.get("generated_at")), fallback=datetime.now().strftime("%Y.%m.%d"))
    application_until = parse_date_to_dot(first_value(row.get("requested_until"), row.get("residence_permit_valid_until"), pick(flat, "requested_until", "permit_valid_until")), fallback="2027.12.31")
    agreement_date = parse_date_to_dot(first_value(row.get("agreement_date"), pick(flat, "agreement_date", "contract_date")), fallback=kelt_dot)

    work_addr = first_value(
        row.get("work_location"), row.get("work_address"), pick(flat, "work_location", "work_address", "planned_work_location"),
        default="2454 Iváncsa, Külterület HRSZ 099/048",
    )

    contact_phone = first_value(
        row.get("phone"), row.get("contact_phone"), row.get("applicant_phone"),
        pick(flat, "phone", "contact_phone", "applicant_phone", "mobile_phone"),
        default="+36309560634",
    )
    contact_email = first_value(
        row.get("email"), row.get("contact_email"), row.get("applicant_email"),
        pick(flat, "email", "contact_email", "applicant_email"),
        default="hatosag@hddirekt.com",
    )

    employer_address_raw = first_value(
        row.get("employer_address"), row.get("employer_registered_address"),
        pick(flat, "employer_address", "employer_registered_address", "registered_office", "employer.seat_address"),
        default="",
    )

    def _split_address_for_employer(addr: str) -> Tuple[str, str, str, str, str, str, str, str]:
        s2 = one_line(addr).strip(" .")
        m2 = re.search(r"^(\d{4})\s+([^,]+),?\s+(.+?)\s+(utca|út|útja|tér|körút|köz|sor|fasor|dűlő|park|u\.)\s+(.+?)(?:\.|$)", s2, re.I)
        if m2:
            z, city, street, stype, house = [x.strip().rstrip(".") for x in m2.groups()]
            return z, city, street, work_street_type(stype), house, "", "", ""
        return "", "", "", "", "", "", "", ""

    emp_z0, emp_c0, emp_s0, emp_t0, emp_h0, emp_b0, emp_f0, emp_d0 = _split_address_for_employer(employer_address_raw)
    if company_code == "HDD":
        emp_defaults = ("1087", "Budapest", "Baross", "tér", "1", "", "3", "14")
    elif company_code == "HRD":
        emp_defaults = ("1095", "Budapest", "Soroksári", "út", "48-54", "", "", "")
    else:
        emp_defaults = ("", "", "", "", "", "", "", "")

    acc = first_value(row.get("accommodation_address"), row.get("planned_accommodation"), pick(flat, "accommodation_address", "planned_accommodation"), default="1087 Budapest, Baross tér 1.")
    # Basic split only for generator text. The v8 extractor will parse from the label text too.
    m = re.search(r"^(\d{4})\s+([^,]+),\s+(.+?)\s+(utca|út|tér|körút|köz|sor|útja|u\.)\s+(.+?)\.?$", acc, re.I)
    hu_zip, hu_city, hu_street, hu_stype, hu_house = ("1087", "Budapest", "Baross", "tér", "1")
    if m:
        hu_zip, hu_city, hu_street, hu_stype, hu_house = [x.strip().rstrip(".") for x in m.groups()]

    return GeneratedPerson(
        app_type=app_type,
        full_name=full_name,
        family_name=family,
        given_name=given,
        birth_family_name=first_value(pick(flat, "birth_family_name"), family),
        birth_given_name=first_value(pick(flat, "birth_given_name"), given),
        gender=first_value(row.get("gender"), pick(flat, "gender"), default="Férfi"),
        birth_place=first_value(row.get("birth_place"), pick(flat, "birth_place", "place_of_birth"), default=""),
        birth_country=birth_country,
        birth_date_dot=parse_date_to_dot(first_value(row.get("birth_date"), pick(flat, "birth_date", "date_of_birth")), fallback="1990.01.01"),
        nationality=nationality,
        mother_family=first_value(pick(flat, "mother_family", "mother_last_name"), default=""),
        mother_given=first_value(pick(flat, "mother_given", "mother_first_name"), default=""),
        passport_no=first_value(row.get("passport_number"), pick(flat, "passport_number", "passport_no"), default="UNKNOWN"),
        passport_issue_place=first_value(pick(flat, "passport_issue_place"), default=""),
        passport_issue_date_dot=parse_date_to_dot(first_value(pick(flat, "passport_issue_date")), fallback="2024.01.01"),
        passport_valid_until_dot=parse_date_to_dot(first_value(row.get("passport_expiry_date"), pick(flat, "passport_expiry_date", "passport_valid_until")), fallback="2034.01.01"),
        foreign_country=normalize_country_from_nationality(nationality),
        foreign_city=first_value(pick(flat, "foreign_city", "address_city"), default=""),
        foreign_street=first_value(pick(flat, "foreign_street", "address_street"), default=""),
        hu_zip=hu_zip,
        hu_city=hu_city,
        hu_street=hu_street,
        hu_stype=hu_stype,
        hu_house=hu_house,
        application_until_dot=application_until,
        kelt_dot=kelt_dot,
        agreement_date_dot=agreement_date,
        feor_code=feor_code,
        feor_name=feor_name,
        employer_name=employer_name,
        company_code=company_code,
        employer_short=employer_short,
        tax_no=first_value(pick(flat, "tax_no", "employer_tax_number"), default="25473480-2-42" if company_code == "HDD" else "24987048-2-43"),
        ksh=first_value(pick(flat, "ksh"), default="25473480-7820-113-01" if company_code == "HDD" else "24987048-7820-113-01"),
        teaor=first_value(pick(flat, "teaor"), default="7820"),
        salary=money_digits(first_value(row.get("salary"), row.get("gross_salary"), pick(flat, "salary", "gross_salary")), fallback="455976"),
        work_addr=work_addr,
        language=first_value(pick(flat, "language", "mother_tongue"), default="ukrán" if "ukrán" in norm(nationality) else "angol"),
        reg_date_words=first_value(pick(flat, "registered_at_words"), default="2024 év 3 hó 11 nap"),
        reg_no=first_value(pick(flat, "registration_number"), default="BP/0702/00010-4/2024"),
        eh_case_id=first_value(input_payload.get("eh_case_id"), row.get("eh_case_id"), pick(flat, "eh_case_id"), default=""),
        package_id=first_value(input_payload.get("package_id"), row.get("package_id"), pick(flat, "package_id"), default=""),
        workflow_case_id=first_value(input_payload.get("workflow_case_id"), row.get("workflow_case_id"), pick(flat, "workflow_case_id"), default=""),
        contact_phone=contact_phone,
        contact_email=contact_email,
        employer_zip=first_value(row.get("employer_zip"), pick(flat, "employer_zip", "employer_postal_code"), emp_z0, emp_defaults[0]),
        employer_city=first_value(row.get("employer_city"), pick(flat, "employer_city", "employer_settlement"), emp_c0, emp_defaults[1]),
        employer_street=first_value(row.get("employer_street"), pick(flat, "employer_street", "employer_street_name"), emp_s0, emp_defaults[2]),
        employer_stype=first_value(row.get("employer_street_type"), row.get("employer_stype"), pick(flat, "employer_street_type", "employer_stype"), emp_t0, emp_defaults[3]),
        employer_house=first_value(row.get("employer_house"), row.get("employer_house_number"), pick(flat, "employer_house", "employer_house_number"), emp_h0, emp_defaults[4]),
        employer_building=first_value(row.get("employer_building"), pick(flat, "employer_building"), emp_b0, emp_defaults[5]),
        employer_staircase=first_value(row.get("employer_staircase"), row.get("employer_stairs"), pick(flat, "employer_staircase", "employer_stairs"), default=""),
        employer_floor=first_value(row.get("employer_floor"), pick(flat, "employer_floor"), emp_f0, emp_defaults[6]),
        employer_door=first_value(row.get("employer_door"), pick(flat, "employer_door"), emp_d0, emp_defaults[7]),
        source_payload=input_payload,
    )

# -----------------------------
# XLSM row mapping
# -----------------------------

def build_xlsm_row(p: GeneratedPerson) -> Dict[str, str]:
    row: Dict[str, str] = {k: "" for k in LETTERS}
    row.update({
        "A": p.kelt_dot,
        "B": p.company_code,
        "C": p.full_name,
        "D": p.family_name,
        "E": p.given_name,
        "F": p.birth_family_name,
        "G": p.gender,
        "H": p.birth_place,
        "I": p.birth_country,
        "J": p.birth_date_dot,
        "K": p.nationality,
        "L": p.mother_family,
        "M": p.mother_given,
        "N": p.passport_no,
        "P": p.foreign_country,
        "Q": p.foreign_city,
        "R": p.foreign_street,
        "X": "Nőtlen/hajadon",
        "Y": "Középfokú",
        "Z": "Szakközépiskola",
        "AA": "",
        "AB": "Nem",
        "AC": "",
        "AD": p.language,
        "AE": p.hu_zip,
        "AF": p.hu_city,
        "AG": p.hu_street,
        "AH": p.hu_stype,
        "AI": p.hu_house,
        "AO": "Bérlő",
        "AQ": p.application_until_dot if p.app_type == "nemzeti" else "Határozatlan ideig",
        "AR": "Foglalkoztatási Jogviszony Alapján",
        "AV": "Magán",
        "AW": p.passport_issue_place,
        "AX": p.passport_issue_date_dot,
        "AY": p.passport_valid_until_dot,
        "AZ": p.agreement_date_dot,
        "BA": p.feor_code,
        "BB": p.feor_name,
        "BD": p.foreign_country or p.birth_country,
        "BF": p.application_until_dot,
        "BG": p.employer_short,
        "BH": p.teaor,
        "BI": p.ksh,
        "BJ": p.tax_no,
        "BK": "01 09 931447" if p.company_code == "HDD" else "01 09 182281",
        "BL": "Kft.",
        "BM": "1087" if p.company_code == "HDD" else "1095",
        "BN": "Baross" if p.company_code == "HDD" else "Soroksári",
        "BO": "tér" if p.company_code == "HDD" else "út",
        "BP": "Tér" if p.company_code == "HDD" else "Út",
        "BQ": "1" if p.company_code == "HDD" else "48-54",
        "BR": "",
        "BS": "03" if p.company_code == "HDD" else "",
        "BT": "14" if p.company_code == "HDD" else "",
        "CO": "Posta",
        "CP": "hatosag@hddirekt.com",
        "CQ": "+36309560634",
        "CR": "foglalkoztató" if p.app_type == "nemzeti" else "minősített kölcsönbeadó",
        "CT": p.salary,
        "CU": "Forint",
        "CV": "az állampolgárságom szerinti állam",
        "CY": "minősített munkaerő kölcsönző",
        "CZ": p.reg_date_words.replace("év", "év").replace("hó", "hó").replace("nap", "nap"),
        "DA": p.reg_no,
        "DB": "1" if p.app_type == "vendeg" else "",
        "DE": "1" if p.app_type == "vendeg" else "",
        "DF": "36" if p.app_type == "nemzeti" else "",
    })
    # Work address split for workbook fields BU:BY. Keep fallback robust.
    wz, wc, ws, wt, wh = split_hu_work_addr(p.work_addr)
    row.update({"BU": wz, "BV": wc, "BW": ws, "BX": wt, "BY": wh})
    return row


def split_hu_work_addr(addr: str) -> Tuple[str, str, str, str, str]:
    s = one_line(addr).strip(" .")
    m_hrsz = re.search(r"^(\d{4})\s+([^,]+),\s*(.*?)\s+HRSZ\s+([0-9A-Za-zÁÉÍÓÖŐÚÜŰáéíóöőúüű/_\-]+)\.?$", s, re.I)
    if m_hrsz:
        z, city, area, hrsz = m_hrsz.groups()
        return z, city, area or "Külterület", "Kültelek", hrsz.rstrip(".")
    m = re.search(r"^(\d{4})\s+([^,]+),\s+(.+?)\s+(utca|út|útja|tér|körút|köz|sor|fasor|dűlő|park|u\.)\s+(.+?)\.?$", s, re.I)
    if m:
        z, city, street, stype, house = m.groups()
        return z, city, street, work_street_type(stype), house.rstrip(".")
    return "", "", "", "", ""


def work_street_type(value: str) -> str:
    m = {
        "utca": "Utca", "u.": "Utca", "út": "Út", "ut": "Út", "útja": "Útja", "utja": "Útja",
        "tér": "Tér", "ter": "Tér", "körút": "Körút", "korut": "Körút", "köz": "Köz", "koz": "Köz",
        "sor": "Sor", "fasor": "Fasor", "dűlő": "Dűlő", "dulo": "Dűlő", "kultelek": "Kültelek", "kültelek": "Kültelek",
    }
    return m.get(norm(value), value[:1].upper() + value[1:] if value else "")

# -----------------------------
# PDF generation
# -----------------------------

# -----------------------------
# Official DOCX/PDF template stamping v3
# -----------------------------

TEMPLATE_MODE_VERSION = "v3.7-official-template-employment96-guest97"


def date_parts_dot(value: str) -> Tuple[str, str, str]:
    d = parse_date_to_dot(value, fallback="")
    if not d:
        return "", "", ""
    y, m, day = d.split(".")
    return y, m, day


def _font_file_for_pdf_stamp() -> Optional[str]:
    candidates = [
        os.environ.get("OIF_EH_PDF_FONT", ""),
        "C:/Windows/Fonts/arial.ttf",
        "C:/Windows/Fonts/Arial.ttf",
        "C:/Windows/Fonts/calibri.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf",
    ]
    for item in candidates:
        if item and Path(item).exists():
            return str(Path(item))
    return None


def _template_search_dirs() -> List[Path]:
    out: List[Path] = []
    env_dir = one_line(os.environ.get("OIF_EH_TEMPLATE_DIR"))
    if env_dir:
        out.append(Path(env_dir))
    here = Path(__file__).resolve().parent
    out.extend([
        here / "templates_pdf",
        here / "templates",
        Path.cwd() / "templates_pdf",
        Path.cwd() / "templates",
        Path("C:/EH_AGENT/oif_templates"),
        Path("C:/EH_AGENT/oif_templates/templates_pdf"),
    ])
    # de-dup while preserving order
    seen = set()
    uniq = []
    for d in out:
        key = str(d).lower()
        if key not in seen:
            seen.add(key)
            uniq.append(d)
    return uniq


def _find_template_file(stem: str, suffixes: Iterable[str]) -> Optional[Path]:
    names = []
    for ext in suffixes:
        names.extend([
            f"{stem}{ext}",
            f"{stem.replace('_', ' ')}{ext}",
        ])
    for d in _template_search_dirs():
        if not d.exists():
            continue
        for name in names:
            cand = d / name
            if cand.exists():
                return cand
        # fuzzy fallback
        for cand in d.glob("*"):
            if cand.suffix.lower() in {x.lower() for x in suffixes}:
                if all(part in norm(cand.stem) for part in norm(stem).split("_") if part):
                    return cand
    return None


def _convert_docx_to_pdf_if_needed(stem: str) -> Optional[Path]:
    pdf = _find_template_file(stem, [".pdf"])
    if pdf:
        return pdf
    docx = _find_template_file(stem, [".docx"])
    if not docx:
        return None

    out_dir = Path(__file__).resolve().parent / "templates_pdf"
    ensure_dir(out_dir)
    expected = out_dir / f"{stem}.pdf"
    if expected.exists():
        return expected

    exe_candidates = [
        os.environ.get("LIBREOFFICE_EXE", ""),
        "soffice",
        "libreoffice",
        "C:/Program Files/LibreOffice/program/soffice.exe",
        "C:/Program Files (x86)/LibreOffice/program/soffice.exe",
    ]
    import subprocess
    last_error = ""
    for exe in exe_candidates:
        if not exe:
            continue
        try:
            cmd = [exe, "--headless", "--convert-to", "pdf", "--outdir", str(out_dir), str(docx)]
            proc = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
            if proc.returncode == 0:
                generated = out_dir / (docx.stem + ".pdf")
                if generated.exists():
                    if generated != expected:
                        try:
                            shutil.copyfile(generated, expected)
                        except Exception:
                            return generated
                    return expected if expected.exists() else generated
            last_error = (proc.stderr or proc.stdout or "")[-500:]
        except Exception as exc:
            last_error = f"{type(exc).__name__}: {exc}"
            continue
    raise RuntimeError(f"DOCX sablon PDF konverzió sikertelen: {docx} | {last_error}")


def _text_kwargs(size: float = 8.2, bold: bool = False) -> Dict[str, Any]:
    font_file = _font_file_for_pdf_stamp()
    if font_file:
        return {"fontsize": size, "fontname": "OIFUnicode", "fontfile": font_file, "color": (0, 0, 0)}
    return {"fontsize": size, "fontname": "helv", "color": (0, 0, 0)}


def _stamp_text(page: Any, x: float, y: float, text: Any, width: float = 180, height: float = 16, size: float = 8.2) -> None:
    s = one_line(text)
    if not s:
        return
    try:
        page.insert_textbox(fitz.Rect(x, y, x + width, y + height), s, **_text_kwargs(size=size))
    except Exception:
        # Last resort with Helvetica and ASCII-ish fallback. We prefer not to fail whole generation on a glyph issue.
        safe = s.encode("latin-1", "replace").decode("latin-1")
        page.insert_textbox(fitz.Rect(x, y, x + width, y + height), safe, fontsize=size, fontname="helv", color=(0, 0, 0))


def _stamp_check(page: Any, x: float, y: float, checked: bool = True) -> None:
    """Legacy text-X checkbox stamp.

    Older main-template coordinates were tuned as text baselines, so this helper
    is kept unchanged for the main 9. melléklet. 9.6/9.7 annex checkboxes now
    use _stamp_check_box with exact rectangle top-left coordinates from the
    official templates.
    """
    if not checked:
        return
    try:
        page.insert_text((x + 1.2, y - 0.6), "X", fontsize=7.8, fontname="helv", color=(0, 0, 0))
    except Exception:
        pass


def _stamp_check_box(page: Any, x0: float, y0: float, checked: bool = True, box: float = 9.2) -> None:
    """Draw a centered X inside a checkbox rectangle.

    Use this for official 9.6 / 9.7 annexes where the checkbox squares are not
    always aligned on the same row/column. The arguments are the actual top-left
    coordinates of the checkbox square in PyMuPDF's top-left coordinate system.
    Drawing two diagonal strokes avoids font-baseline drift and keeps the mark
    inside the box even when checkboxes sit at different y positions.
    """
    if not checked:
        return
    try:
        inset = 1.7
        page.draw_line((x0 + inset, y0 + inset), (x0 + box - inset, y0 + box - inset), color=(0, 0, 0), width=0.75)
        page.draw_line((x0 + box - inset, y0 + inset), (x0 + inset, y0 + box - inset), color=(0, 0, 0), width=0.75)
    except Exception:
        pass


def _stamp_date_parts(page: Any, x_year: float, x_month: float, x_day: float, y: float, value: str, size: float = 8.2) -> None:
    yy, mm_, dd = date_parts_dot(value)
    _stamp_text(page, x_year, y, yy, 34, 13, size)
    _stamp_text(page, x_month, y, mm_, 34, 15, size)
    _stamp_text(page, x_day, y, dd, 34, 15, size)


def _employer_address_parts(p: GeneratedPerson) -> Tuple[str, str, str, str, str, str, str, str]:
    if any([p.employer_zip, p.employer_city, p.employer_street, p.employer_house]):
        return (
            p.employer_zip, p.employer_city, p.employer_street, p.employer_stype,
            p.employer_house, p.employer_building, p.employer_floor, p.employer_door,
        )
    if p.company_code == "HRD":
        return "1095", "Budapest", "Soroksári", "út", "48-54", "", "", ""
    return "1087", "Budapest", "Baross", "tér", "1", "", "3", "14"


def stamp_main_application_template(p: GeneratedPerson, template_pdf: Path, out_pdf: Path) -> Path:
    doc = fitz.open(str(template_pdf))

    # Page 1 - application header + personal name
    page = doc[0]
    # Extension is default for current OIF hosszabbítás workflow; first issue can be forced by process_kind.
    process_hay = norm(json.dumps(p.source_payload, ensure_ascii=False)[:2000])
    is_first_issue = any(x in process_hay for x in ["first_issue", "elso alkalom", "első alkalom", "new_permit"])
    _stamp_check(page, 32.8, 617.5, checked=is_first_issue)
    _stamp_check(page, 32.2, 633.1, checked=not is_first_issue)
    _stamp_text(page, 92, 645, "+36309560634", 160, 13, 8.4)
    _stamp_text(page, 356, 645, "hatosag@hddirekt.com", 190, 13, 8.4)
    _stamp_check(page, 32.2, 714.0, True)  # átveszi a kiállító hatóságnál
    _stamp_text(page, 150, 746, p.family_name, 150, 14, 8.6)
    _stamp_text(page, 416, 746, p.given_name, 150, 14, 8.6)

    # Page 2 - personal/passport/accommodation/insurance
    page = doc[1]
    _stamp_text(page, 122, 73, p.birth_family_name, 155, 13, 8.2)
    _stamp_text(page, 388, 73, p.birth_given_name, 150, 13, 8.2)
    _stamp_text(page, 148, 91, p.mother_family, 155, 13, 8.2)
    _stamp_text(page, 414, 91, p.mother_given, 150, 13, 8.2)
    _stamp_check(page, 62.9, 119.1, checked=norm(p.gender).startswith("fer") or norm(p.gender).startswith("male"))
    _stamp_check(page, 96.3, 119.1, checked=norm(p.gender).startswith("no") or norm(p.gender).startswith("female"))
    _stamp_date_parts(page, 87, 126, 166, 126, p.birth_date_dot, 8.2)
    _stamp_text(page, 300, 126, p.birth_place, 125, 13, 8.2)
    _stamp_text(page, 477, 126, p.birth_country, 90, 13, 8.2)
    _stamp_text(page, 106, 143, p.nationality, 155, 13, 8.2)
    _stamp_text(page, 105, 161, "", 80, 13, 8.2)
    _stamp_check(page, 307.8, 185.1, True)  # középfokú
    _stamp_text(page, 445, 188, "", 125, 13, 8.2)

    _stamp_text(page, 95, 225, p.passport_no, 120, 13, 8.2)
    _stamp_date_parts(page, 338, 380, 422, 225, p.passport_issue_date_dot, 8.2)
    _stamp_text(page, 468, 225, p.passport_issue_place, 95, 13, 8.2)
    _stamp_check(page, 99.0, 256.1, True)  # magánútlevél
    _stamp_date_parts(page, 426, 468, 510, 246, p.passport_valid_until_dot, 8.2)

    _stamp_text(page, 194, 283, p.hu_zip, 35, 13, 8.2)
    _stamp_text(page, 274, 283, p.hu_city, 115, 13, 8.2)
    _stamp_text(page, 463, 283, p.hu_street, 105, 13, 8.2)
    _stamp_text(page, 107, 308, p.hu_stype, 45, 13, 8.2)
    _stamp_text(page, 195, 308, p.hu_house, 34, 13, 8.2)
    _stamp_check(page, 248.1, 343.1, True)  # bérlő
    _stamp_check(page, 37.4, 398.0, True)   # foglalkoztatási jogviszony alapján

    # Page 3 - foreign address + default no/no questions
    page = doc[2]
    # 7. Egyéb adatok / állandó vagy szokásos tartózkodási helye
    _stamp_text(page, 75, 475, p.foreign_country, 115, 13, 8.0)
    _stamp_text(page, 245, 475, p.foreign_city, 120, 13, 8.0)
    _stamp_text(page, 482, 475, p.foreign_street, 90, 13, 8.0)

    # Default answers: nem. Coordinates are tuned to the official Word/PDF template checkboxes.
    for x, y in [
        (427, 500),  # más schengeni okmány? nem
        (66, 567),   # korábban elutasított kérelem? nem
        (66, 593),   # korábban büntetve? nem
        (66, 619),   # kiutasították-e? nem
        (66, 672),   # fertőző betegség / hordozó? nem
        (66, 709),   # kötelező/rendszeres egészségügyi ellátás? nem
        (67, 740),   # kiskorú gyermek együtt utazik? nem
    ]:
        _stamp_check(page, x, y, True)

    # Page 4 - requested until, purpose and date/sign blocks
    page = doc[3]
    _stamp_date_parts(page, 242, 287, 327, 90, p.application_until_dot, 8.3)
    # Mark Nemzeti Kártya 9.12 or vendeg 9.7 if this template reused.
    if p.app_type == "nemzeti":
        _stamp_check(page, 35, 276, True)
    elif p.app_type == "foglalkoztatasi":
        _stamp_check(page, 35, 188, True)
    elif p.app_type == "vendeg":
        _stamp_check(page, 35, 203, True)
    # A "csatolt betétlap(ok)" sorba nem írunk külön "9.12" feliratot: a cél jelölése fent checkboxszal történik.
    # Kelt / aláírás mezőket egyelőre üresen hagyjuk, hogy ne csússzon rá a hivatalos nyilatkozati szövegre.

    # Page 5 - country of return + date
    page = doc[4]
    _stamp_check(page, 38.4, 128.1, True)  # az állampolgárságom szerinti állam
    # Hatósági / aláírási részt nem töltünk automatikusan.

    ensure_dir(out_pdf.parent)
    doc.save(str(out_pdf), garbage=4, deflate=True)
    doc.close()
    return out_pdf


def stamp_nemzeti_annex_template(p: GeneratedPerson, template_pdf: Path, out_pdf: Path) -> Path:
    doc = fitz.open(str(template_pdf))

    page = doc[0]
    _stamp_check(page, 53, 370, True)  # foglalkoztató útján
    _stamp_text(page, 110, 379, "+36309560634", 150, 13, 8.4)
    _stamp_text(page, 103, 406, "hatosag@hddirekt.com", 190, 13, 8.4)
    employer_addr = "1087 Budapest, Baross tér 1. 3. em. 14." if p.company_code == "HDD" else "1095 Budapest, Soroksári út 48-54."
    _stamp_text(page, 50, 490, employer_addr, 260, 16, 8.1)
    _stamp_text(page, 50, 533, employer_addr, 260, 16, 8.1)
    _stamp_check(page, 53, 678, True)  # okmány postai úton kéri
    _stamp_text(page, 52, 738, p.salary + " HUF", 190, 14, 8.3)

    page = doc[1]
    zip_, city, street, stype, house, building, floor, door = _employer_address_parts(p)
    _stamp_text(page, 66, 149, p.employer_name, 285, 13, 8.2)
    _stamp_text(page, 104, 195, zip_, 54, 13, 8.2)
    _stamp_text(page, 217, 195, city, 118, 13, 8.2)
    _stamp_text(page, 420, 195, street, 114, 13, 8.2)
    _stamp_text(page, 124, 220, stype, 38, 13, 8.2)
    _stamp_text(page, 202, 220, house, 50, 13, 8.2)
    _stamp_text(page, 284, 220, building, 35, 13, 8.2)
    _stamp_text(page, 443, 220, floor, 40, 13, 8.2)
    _stamp_text(page, 524, 220, door, 35, 13, 8.2)
    _stamp_text(page, 217, 245, p.tax_no, 72, 13, 8.2)
    _stamp_text(page, 339, 245, p.ksh, 66, 13, 8.2)
    _stamp_text(page, 474, 245, p.teaor, 52, 13, 8.2)
    _stamp_text(page, 46, 303, "-", 90, 13, 8.2)
    _stamp_check(page, 219.4, 334.1, True)  # szakközépiskola
    _stamp_text(page, 394, 303, "", 110, 13, 8.2)

    _stamp_check(page, 53.4, 446.1, True)  # egyetlen munkavégzési hely igen
    _stamp_text(page, 83, 457, p.work_addr, 120, 42, 7.6)
    _stamp_check(page, 260.4, 447.1, True)  # több vármegye: nem
    _stamp_check(page, 462.4, 447.1, True)  # több telephely: nem

    _stamp_date_parts(page, 47, 86, 126, 545, p.agreement_date_dot, 8.2)
    _stamp_text(page, 323, 544, p.feor_code, 110, 13, 8.2)
    _stamp_text(page, 298, 583, "0 év", 90, 13, 8.2)
    _stamp_text(page, 322, 598, "", 130, 13, 8.2)
    _stamp_text(page, 100, 625, p.language, 100, 13, 8.2)
    _stamp_text(page, 136, 638, "angol", 90, 13, 8.2)
    _stamp_check(page, 164.4, 660.1, True)  # beszél magyarul? nem
    _stamp_check(page, 84.4, 691.1, True)   # korábban dolgozott Mo-on? nem

    page = doc[2]
    # 10. pont: távozási kötelezettség célországa
    _stamp_text(page, 423, 124, p.foreign_country or p.birth_country, 72, 13, 8.0)
    _stamp_check(page, 52.4, 226.1, True)  # az állampolgárságom szerinti állam
    _stamp_check(page, 54.4, 366.1, True)  # 242 § fennáll? Nem
    _stamp_check(page, 52.4, 435.1, True)  # munkavállalási engedély alól mentes? Nem

    ensure_dir(out_pdf.parent)
    doc.save(str(out_pdf), garbage=4, deflate=True)
    doc.close()
    return out_pdf



def stamp_foglalkoztatasi_annex_template(p: GeneratedPerson, template_pdf: Path, out_pdf: Path) -> Path:
    """Stamp the official 9.6 Foglalkoztatás annex template.

    Coordinates are tuned against the official empty PDF template supplied by the user.
    The default answers follow the filled example package:
      - okmány postai úton,
      - belföldi foglalkoztatóval fennálló munkaviszony,
      - one work location,
      - no multi-county / no multi-site,
      - no Hungarian language,
      - no previous HU employment,
      - no exemptions in points 11-13.
    """
    doc = fitz.open(str(template_pdf))

    page = doc[0]
    # Header: document delivery/contact
    _stamp_check_box(page, 51.0, 308.5, True)  # postai úton kéri
    _stamp_text(page, 379, 296, p.contact_phone, 150, 13, 8.2)
    _stamp_text(page, 379, 311, p.contact_email, 180, 13, 8.2)

    # 1. Employment basis
    _stamp_check_box(page, 46.7, 341.8, True)  # belföldi foglalkoztatóval fennálló munkaviszony

    # 2. Livelihood data
    _stamp_text(page, 48, 444, f"{p.salary} /HÓ", 125, 13, 8.1)

    # 3. Hungarian employer data
    zip_, city, street, stype, house, building, floor, door = _employer_address_parts(p)
    staircase = p.employer_staircase
    _stamp_text(page, 64, 532, p.employer_name, 360, 13, 8.2)
    _stamp_text(page, 107, 568, zip_, 55, 13, 8.2)
    _stamp_text(page, 216, 568, city, 118, 13, 8.2)
    _stamp_text(page, 440, 568, street, 110, 13, 8.2)
    _stamp_text(page, 50, 603, stype, 68, 13, 8.0)
    _stamp_text(page, 181, 603, house, 58, 13, 8.2)
    _stamp_text(page, 258, 603, building, 44, 13, 8.2)
    _stamp_text(page, 332, 603, staircase, 50, 13, 8.2)
    _stamp_text(page, 434, 603, floor, 42, 13, 8.2)
    _stamp_text(page, 520, 603, door, 38, 13, 8.2)
    _stamp_text(page, 215, 617, p.tax_no, 70, 13, 6.2)
    _stamp_text(page, 335, 617, p.ksh, 72, 13, 5.8)
    _stamp_text(page, 488, 617, p.teaor, 45, 13, 6.6)

    # 4-6 Qualification / education / former occupation
    _stamp_text(page, 47, 676, "Nincs", 110, 13, 8.2)
    _stamp_check_box(page, 216.5, 672.8, True)  # szakközépiskola
    _stamp_text(page, 394, 676, "Nincs", 120, 13, 8.2)

    # 7. Work location(s)
    _stamp_check_box(page, 51.4, 734.3, True)  # Egyetlen munkavégzési hely: igen
    _stamp_text(page, 49, 763, p.work_addr, 154, 44, 7.5)
    _stamp_check_box(page, 255.3, 742.8, True)  # több vármegye: nem
    _stamp_check_box(page, 451.7, 757.3, True)  # több telephely: nem

    page = doc[1]
    # 8-9 Agreement date and FEOR
    _stamp_date_parts(page, 55, 93, 136, 61, p.agreement_date_dot, 8.2)
    _stamp_text(page, 402, 69, f"{p.feor_code} {p.feor_name}".strip(), 150, 13, 8.0)

    # 10. Skills/language/prior employment
    # szakmai gyakorlati idő: üresen hagyva, mert nincs megbízható forrásadat
    _stamp_text(page, 104, 143, p.language, 120, 13, 8.2)
    _stamp_text(page, 119, 156, "", 120, 13, 8.2)
    _stamp_check_box(page, 160.7, 167.2, True)  # Beszél magyarul? nem
    _stamp_check_box(page, 253.6, 193.1, True)  # Korábban dolgozott Magyarországon? nem

    # 11-13. Default: Nem.
    _stamp_check_box(page, 50.9, 311.2, True)  # 11 Nem
    _stamp_check_box(page, 48.8, 379.2, True)  # 12 Nem
    _stamp_check_box(page, 46.6, 454.9, True)  # 13 Nem

    ensure_dir(out_pdf.parent)
    doc.save(str(out_pdf), garbage=4, deflate=True)
    doc.close()
    return out_pdf


def generate_foglalkoztatasi_official_template_pdf(p: GeneratedPerson, out_pdf: Path, include_qr_backpage: bool = True) -> Path:
    if fitz is None:
        raise RuntimeError("PyMuPDF/fitz hiányzik, official template stamping nem használható")

    main_template = _convert_docx_to_pdf_if_needed("9_tartozkodasi_engedely_kerelem_hu")
    annex_template = _convert_docx_to_pdf_if_needed("9_6_foglalkoztatas_hu")
    if not main_template or not annex_template:
        raise FileNotFoundError("Hiányzó official template PDF/DOCX: 9_tartozkodasi_engedely_kerelem_hu vagy 9_6_foglalkoztatas_hu")

    ensure_dir(out_pdf.parent)
    work = out_pdf.parent / "_template_work"
    ensure_dir(work)
    main_out = work / f"main_9_{uuid.uuid4().hex[:8]}.pdf"
    annex_out = work / f"annex_9_6_{uuid.uuid4().hex[:8]}.pdf"

    stamp_main_application_template(p, main_template, main_out)
    stamp_foglalkoztatasi_annex_template(p, annex_template, annex_out)

    merged = fitz.open()
    d1 = fitz.open(str(main_out))
    d2 = fitz.open(str(annex_out))
    merged.insert_pdf(d1)
    merged.insert_pdf(d2)
    d1.close()
    d2.close()
    if include_qr_backpage:
        _append_qr_page_to_doc(merged, p)
    merged.save(str(out_pdf), garbage=4, deflate=True)
    merged.close()
    try:
        shutil.rmtree(work)
    except Exception:
        pass
    return out_pdf


def _stamp_date_parts_flexible(page: Any, x_year: float, x_month: float, x_day: float, y: float, value: str, size: float = 8.2) -> None:
    """Stamp a date that may arrive as YYYY.MM.DD or as '2024 év 03 hó 11 nap'."""
    nums = re.findall(r"\d+", one_line(value))
    if len(nums) >= 3:
        yy = f"{int(nums[0]):04d}"
        mm_ = f"{int(nums[1]):02d}"
        dd = f"{int(nums[2]):02d}"
        _stamp_text(page, x_year, y, yy, 34, 13, size)
        _stamp_text(page, x_month, y, mm_, 25, 13, size)
        _stamp_text(page, x_day, y, dd, 25, 13, size)


def _format_salary_hu(value: str) -> str:
    digits = re.sub(r"\D", "", one_line(value)) or "455976"
    try:
        return f"{int(digits):,}".replace(",", " ") + " FT/HÓ"
    except Exception:
        return f"{digits} FT/HÓ"


def stamp_vendeg_annex_template(p: GeneratedPerson, template_pdf: Path, out_pdf: Path) -> Path:
    """Stamp the official 9.7 Vendégmunkás-tartózkodási engedély annex template.

    Default answers follow the user-supplied Imperal FEOR-mód filled package:
      - submission by minősített kölcsönbeadó,
      - qualified temporary work agency employer type,
      - Btátv. 34 registry: igen with registration date/number,
      - one work location,
      - no multi-county / no multi-site,
      - no Hungarian language,
      - no previous HU employment,
      - point 12: Igen, 242. § (7) 1. pont,
      - point 13: Nem.
    """
    doc = fitz.open(str(template_pdf))

    # Page 1 - submission/contact, livelihood, employer data
    page = doc[0]
    _stamp_check_box(page, 261.2, 334.1, True)  # A kérelem benyújtása: minősített kölcsönbeadó által
    _stamp_text(page, 415, 352, p.contact_phone, 145, 13, 8.2)
    _stamp_text(page, 382, 375, p.contact_email, 185, 13, 8.2)
    _stamp_text(page, 50, 432, "5000 Szolnok, Kápolna út 3.", 250, 18, 8.2)
    # A foglalkoztató székhelyét külön nem ismételjük itt, lent részletes munkáltatói adatként szerepel.

    _stamp_text(page, 50, 604, _format_salary_hu(p.salary), 180, 13, 8.2)

    zip_, city, street, stype, house, building, floor, door = _employer_address_parts(p)
    staircase = p.employer_staircase
    _stamp_text(page, 66, 689, p.employer_name.upper(), 365, 13, 8.0)
    _stamp_text(page, 105, 725, zip_, 48, 13, 8.0)
    _stamp_text(page, 208, 725, city, 120, 13, 8.0)
    _stamp_text(page, 420, 725, street, 120, 13, 8.0)
    _stamp_text(page, 108, 758, stype, 32, 13, 7.8)
    _stamp_text(page, 180, 746, house, 54, 13, 8.0)
    _stamp_text(page, 277, 746, building, 36, 13, 8.0)
    _stamp_text(page, 370, 746, staircase, 36, 13, 8.0)
    _stamp_text(page, 442, 746, floor, 48, 13, 8.0)
    _stamp_text(page, 528, 746, door, 36, 13, 8.0)
    _stamp_text(page, 216, 776, p.tax_no, 70, 13, 6.2)
    _stamp_text(page, 338, 776, p.ksh, 72, 13, 5.8)
    _stamp_text(page, 473, 776, p.teaor, 48, 13, 6.6)

    # Page 2 - employer type, registry, education/work/FEOR/language/default legal points
    page = doc[1]
    _stamp_check_box(page, 44.5, 90.2, True)  # 3. minősített munkaerő kölcsönző
    _stamp_check_box(page, 47.0, 166.4, True)  # 4. szerepel? igen
    _stamp_date_parts_flexible(page, 160, 200, 240, 180, p.reg_date_words or "2024 év 03 hó 11 nap", 8.0)
    _stamp_text(page, 371, 180, p.reg_no, 150, 13, 7.2)

    _stamp_text(page, 48, 224, "Nincs", 110, 13, 8.0)
    _stamp_check_box(page, 217.1, 239.7, True)  # szakközépiskola
    _stamp_text(page, 429, 224, "Nincs", 110, 13, 8.0)

    _stamp_check_box(page, 52.0, 308.6, True)  # egyetlen munkavégzési hely: igen
    _stamp_text(page, 48, 335, p.work_addr, 155, 35, 7.2)
    _stamp_check_box(page, 255.9, 305.6, True)  # több vármegye: nem
    _stamp_check_box(page, 468.5, 317.2, True)  # több telephely: nem

    _stamp_date_parts(page, 47, 86, 126, 369, p.agreement_date_dot, 8.0)
    _stamp_text(page, 323, 369, p.feor_code, 70, 13, 8.0)

    _stamp_text(page, 100, 449, p.language, 100, 13, 8.0)
    _stamp_check_box(page, 161.4, 479.2, True)  # Beszél magyarul? nem
    _stamp_check_box(page, 254.1, 505.3, True)  # Korábban dolgozott Magyarországon? nem

    _stamp_check_box(page, 51.5, 618.9, True)   # 12 Igen
    _stamp_text(page, 267, 618, "1", 15, 13, 8.0)
    _stamp_check_box(page, 51.5, 698.7, True)   # 13 Nem

    ensure_dir(out_pdf.parent)
    doc.save(str(out_pdf), garbage=4, deflate=True)
    doc.close()
    return out_pdf


def generate_vendeg_official_template_pdf(p: GeneratedPerson, out_pdf: Path, include_qr_backpage: bool = True) -> Path:
    if fitz is None:
        raise RuntimeError("PyMuPDF/fitz hiányzik, official template stamping nem használható")

    main_template = _convert_docx_to_pdf_if_needed("9_tartozkodasi_engedely_kerelem_hu")
    annex_template = _convert_docx_to_pdf_if_needed("9_7_vendegmunkas_hu")
    if not main_template or not annex_template:
        raise FileNotFoundError("Hiányzó official template PDF/DOCX: 9_tartozkodasi_engedely_kerelem_hu vagy 9_7_vendegmunkas_hu")

    ensure_dir(out_pdf.parent)
    work = out_pdf.parent / "_template_work"
    ensure_dir(work)
    main_out = work / f"main_9_{uuid.uuid4().hex[:8]}.pdf"
    annex_out = work / f"annex_9_7_{uuid.uuid4().hex[:8]}.pdf"

    stamp_main_application_template(p, main_template, main_out)
    stamp_vendeg_annex_template(p, annex_template, annex_out)

    merged = fitz.open()
    d1 = fitz.open(str(main_out))
    d2 = fitz.open(str(annex_out))
    merged.insert_pdf(d1)
    merged.insert_pdf(d2)
    d1.close()
    d2.close()
    if include_qr_backpage:
        _append_qr_page_to_doc(merged, p)
    merged.save(str(out_pdf), garbage=4, deflate=True)
    merged.close()
    try:
        shutil.rmtree(work)
    except Exception:
        pass
    return out_pdf

def _append_qr_page_to_doc(doc: Any, p: GeneratedPerson) -> None:
    w, h = A4
    page = doc.new_page(width=w, height=h)
    title = "OIF/EH csomagazonosító hátlap"
    try:
        page.insert_text((50, 55), title, fontsize=13, fontname="helv", color=(0, 0, 0))
    except Exception:
        pass
    qr_payload = {
        "v": 3,
        "type": "oif_eh_official_template_pdf",
        "template_mode": TEMPLATE_MODE_VERSION,
        "app_type": p.app_type,
        "package_id": p.package_id,
        "workflow_case_id": p.workflow_case_id,
        "full_name_hash": hashlib.sha256(p.full_name.encode("utf-8")).hexdigest()[:16],
        "generated_at": datetime.now().isoformat(timespec="seconds"),
    }
    qr_path = Path(tempfile.gettempdir()) / f"oif_eh_qr_{uuid.uuid4().hex}.png"
    try:
        make_qr_image(qr_payload, qr_path)
        if qr_path.exists():
            page.insert_image(fitz.Rect(50, 80, 210, 240), filename=str(qr_path))
    except Exception:
        pass
    finally:
        try:
            qr_path.unlink()
        except Exception:
            pass
    y = 270
    for k, v in qr_payload.items():
        _stamp_text(page, 50, y, f"{k}: {v}", 470, 14, 8.0)
        y += 15


def generate_nemzeti_official_template_pdf(p: GeneratedPerson, out_pdf: Path, include_qr_backpage: bool = True) -> Path:
    if fitz is None:
        raise RuntimeError("PyMuPDF/fitz hiányzik, official template stamping nem használható")

    main_template = _convert_docx_to_pdf_if_needed("9_tartozkodasi_engedely_kerelem_hu")
    annex_template = _convert_docx_to_pdf_if_needed("9_12_nemzeti_kartya_hu")
    if not main_template or not annex_template:
        raise FileNotFoundError("Hiányzó official template PDF/DOCX: 9_tartozkodasi_engedely_kerelem_hu vagy 9_12_nemzeti_kartya_hu")

    ensure_dir(out_pdf.parent)
    work = out_pdf.parent / "_template_work"
    ensure_dir(work)
    main_out = work / f"main_9_{uuid.uuid4().hex[:8]}.pdf"
    annex_out = work / f"annex_9_12_{uuid.uuid4().hex[:8]}.pdf"

    stamp_main_application_template(p, main_template, main_out)
    stamp_nemzeti_annex_template(p, annex_template, annex_out)

    merged = fitz.open()
    d1 = fitz.open(str(main_out))
    d2 = fitz.open(str(annex_out))
    merged.insert_pdf(d1)
    merged.insert_pdf(d2)
    d1.close()
    d2.close()
    if include_qr_backpage:
        _append_qr_page_to_doc(merged, p)
    merged.save(str(out_pdf), garbage=4, deflate=True)
    merged.close()
    try:
        shutil.rmtree(work)
    except Exception:
        pass
    return out_pdf


def generate_filled_source_pdf(p: GeneratedPerson, out_pdf: Path, include_qr_backpage: bool = True) -> Path:
    """
    v3: Nemzeti Kártya esetén a hivatalos OIF Word sablonból készült PDF-re stampelünk.
    Ha bármi hiányzik, fallbackel a régi extractor-kompatibilis ReportLab PDF-re.
    """
    force_fallback = str(os.environ.get("OIF_EH_FORCE_FALLBACK_PDF", "")).strip().lower() in {"1", "true", "yes"}
    if p.app_type == "vendeg" and not force_fallback:
        try:
            return generate_vendeg_official_template_pdf(p, out_pdf, include_qr_backpage=include_qr_backpage)
        except Exception as exc:
            try:
                p.source_payload.setdefault("template_generation_warning", f"official_template_failed: {type(exc).__name__}: {exc}")
            except Exception:
                pass
            return _generate_fallback_source_pdf(p, out_pdf, include_qr_backpage=include_qr_backpage)
    if p.app_type == "nemzeti" and not force_fallback:
        try:
            return generate_nemzeti_official_template_pdf(p, out_pdf, include_qr_backpage=include_qr_backpage)
        except Exception as exc:
            # Fallback legyen működésbiztos, de tegyük láthatóvá a manifestben a source_payload-on keresztül.
            try:
                p.source_payload.setdefault("template_generation_warning", f"official_template_failed: {type(exc).__name__}: {exc}")
            except Exception:
                pass
            return _generate_fallback_source_pdf(p, out_pdf, include_qr_backpage=include_qr_backpage)
    if p.app_type == "foglalkoztatasi" and not force_fallback:
        try:
            return generate_foglalkoztatasi_official_template_pdf(p, out_pdf, include_qr_backpage=include_qr_backpage)
        except Exception as exc:
            try:
                p.source_payload.setdefault("template_generation_warning", f"official_template_failed: {type(exc).__name__}: {exc}")
            except Exception:
                pass
            return _generate_fallback_source_pdf(p, out_pdf, include_qr_backpage=include_qr_backpage)
    return _generate_fallback_source_pdf(p, out_pdf, include_qr_backpage=include_qr_backpage)


def draw_wrapped(c: canvas.Canvas, text: str, x: float, y: float, max_width: float, leading: float = 12, font: str = PDF_FONT, size: int = 9) -> float:
    c.setFont(font, size)
    words = text.split()
    line = ""
    for word in words:
        test = (line + " " + word).strip()
        if c.stringWidth(test, font, size) <= max_width or not line:
            line = test
        else:
            c.drawString(x, y, line)
            y -= leading
            line = word
    if line:
        c.drawString(x, y, line)
        y -= leading
    return y


def add_page_header(c: canvas.Canvas, title: str, p: GeneratedPerson, page_no: int) -> float:
    w, h = A4
    c.setFont(PDF_FONT_BOLD, 12)
    c.drawString(18 * mm, h - 18 * mm, title)
    c.setFont(PDF_FONT, 8)
    c.drawRightString(w - 18 * mm, h - 18 * mm, f"{p.full_name} | {p.package_id or p.workflow_case_id or ''} | oldal {page_no}")
    c.line(18 * mm, h - 21 * mm, w - 18 * mm, h - 21 * mm)
    return h - 30 * mm


def make_qr_image(payload: Dict[str, Any], out_path: Path) -> Optional[Path]:
    if qrcode is None:
        return None
    img = qrcode.make(json.dumps(payload, ensure_ascii=False, sort_keys=True))
    img.save(out_path)
    return out_path


def _generate_fallback_source_pdf(p: GeneratedPerson, out_pdf: Path, include_qr_backpage: bool = True) -> Path:
    ensure_dir(out_pdf.parent)
    c = canvas.Canvas(str(out_pdf), pagesize=A4)
    w, h = A4

    # Page 1: main application sections with extractor-compatible labels
    y = add_page_header(c, "9. melléklet - Tartózkodási engedély iránti kérelem", p, 1)
    c.setFont(PDF_FONT_BOLD, 10)
    c.drawString(18 * mm, y, "1. A kérelmező személyes adatai")
    y -= 14
    lines = [
        f"családi név (útlevél szerint): {p.family_name} utónév (útlevél szerint): {p.given_name} születési családi név: {p.birth_family_name} születési utónév: {p.birth_given_name}",
        f"E-mail cím: hatosag@hddirekt.com Telefonszám: +36309560634",
        "",
        "2. A kérelmező útlevelének adatai",
        f"anyja születési családi neve: {p.mother_family} anyja születési utóneve: {p.mother_given} nem: {p.gender}",
        f"születési idő: {p.birth_date_dot} születési hely (település): {p.birth_place} ország: {p.birth_country} állampolgársága: {p.nationality} nemzetisége:",
        f"útlevél száma: {p.passport_no} kiállításának ideje, helye: {p.passport_issue_date_dot}, {p.passport_issue_place} útlevél típusa: magánútlevél érvényességi ideje: {p.passport_valid_until_dot} 3. A kérelmező",
        f"irányítószám: {p.hu_zip} település: {p.hu_city} közterület neve: {p.hu_street} közterület jellege: {p.hu_stype} házszám: {p.hu_house} épület: lépcsőház: szint: ajtó: egyéb:",
        "4. Teljes körű egészségbiztosítás: Foglalkoztatási Jogviszony Alapján",
    ]
    for line in lines:
        y = draw_wrapped(c, line, 18 * mm, y, w - 36 * mm, leading=12, size=9) if line else y - 8
    c.showPage()

    # Page 2: other data. Keep it separate so the legacy extractor does not confuse birth country with foreign address.
    y = add_page_header(c, "7. Egyéb adatok", p, 2)
    lines = [
        "7. Egyéb adatok",
        f"Ország: {p.foreign_country} Település: {p.foreign_city} Közterület neve: {p.foreign_street} Rendelkezik-e Magyarországon szállással: Igen",
        f"Meddig kérelmezi tartózkodása engedélyezését? {p.application_until_dot} Kijelentem, hogy a fenti adatok a valóságnak megfelelnek.",
        f"Kelt: Budapest, {p.kelt_dot}",
    ]
    for line in lines:
        y = draw_wrapped(c, line, 18 * mm, y, w - 36 * mm, leading=12, size=9)
    c.showPage()

    # Page 3: annex start. This is what find_annex_page() needs.
    annex_no = "9.12" if p.app_type == "nemzeti" else "9.7"
    annex_name = "Nemzeti Kártya" if p.app_type == "nemzeti" else "Vendégmunkás-tartózkodási engedély"
    y = add_page_header(c, f"{annex_no}. Betétlap - {annex_name}", p, 3)
    c.setFont(PDF_FONT_BOLD, 11)
    c.drawString(18 * mm, y, f"{annex_no}. Betétlap tartózkodási engedély - BETÉTLAP")
    y -= 16
    lines = [
        "2. Magyarországi munkáltató adatai",
        f"név: {p.employer_name} székhely címe: 1087 Budapest, Baross tér 1.",
        f"Munkáltató adószáma /adóazonosító jele: {p.tax_no} KSH-szám: {p.ksh} TEÁOR száma: {p.teaor} 3.",
        f"Nyilvántartásba vétel napja: {p.reg_date_words}, Nyilvántartási szám: {p.reg_no} 5. Munkakör",
        f"3. Munkakör betöltéséhez szükséges szakképzettsége: 4. Iskolai végzettsége: szakközépiskola",
        f"5. Magyarországra érkezést megelőző foglalkozása: 6. Munkavégzés helye(i):",
        f"Anyanyelve: {p.language} Egyéb nyelvismerete: angol",
        f"munkaviszonyból származó várható jövedelem összege: {p.salary} előző évi",
        f"{ '10' if p.app_type == 'nemzeti' else '8' }. Munkakör (FEOR szám): {p.feor_code}",
        f"címe(i): {p.work_addr} A munka természetéből adódóan több helyen történik-e munkavégzés: Nem",
        f"Foglalkoztatóval kötött előzetes megállapodás kelte: {p.agreement_date_dot} 10. Munkakör megnevezése: {p.feor_name}",
    ]
    if p.app_type == "vendeg":
        lines.append("242. § (7) bekezdés 1. pontja szerinti foglalkoztatás: Igen")
    if p.app_type == "nemzeti":
        lines.append("445. § szerinti Nemzeti Kártya 36. pontja")
    for line in lines:
        y = draw_wrapped(c, line, 18 * mm, y, w - 36 * mm, leading=12, size=9)
    c.showPage()

    # Page 4: pre-employment agreement text. Used by extractor fallbacks.
    y = add_page_header(c, "Előzetes megállapodás foglalkoztatásra / Pre-employment agreement", p, 4)
    lines = [
        "Előzetes megállapodás foglalkoztatásra",
        f"Név: {p.full_name} Születési név: {p.birth_family_name} {p.birth_given_name} Gender: {p.gender}",
        f"Állampolgársága: {p.nationality} mint harmadik országbeli állampolgár",
        f"Útlevél száma: {p.passport_no}",
        f"6. A harmadik országbeli állampolgár által betöltendő munkakör megnevezése: / {p.feor_code} - {p.feor_name} / 6,",
        f"9. A munkavégzés helye(i): {p.work_addr} 10. A munkabér összege: {p.salary} HUF",
        f"Kelt/Date: Budapest, {p.kelt_dot}",
    ]
    for line in lines:
        y = draw_wrapped(c, line, 18 * mm, y, w - 36 * mm, leading=13, size=10)
    c.showPage()

    # Page 5: QR back page / package identifier
    if include_qr_backpage:
        y = add_page_header(c, "OIF/EH csomagazonosító hátlap", p, 5)
        qr_payload = {
            "v": 1,
            "type": "oif_eh_generated_source_pdf",
            "app_type": p.app_type,
            "package_id": p.package_id,
            "workflow_case_id": p.workflow_case_id,
            "full_name_hash": hashlib.sha256(p.full_name.encode("utf-8")).hexdigest()[:16],
            "generated_at": datetime.now().isoformat(timespec="seconds"),
        }
        qr_path = out_pdf.with_suffix(".qr.png")
        make_qr_image(qr_payload, qr_path)
        c.setFont(PDF_FONT_BOLD, 11)
        c.drawString(18 * mm, y, "QR / belső csomagazonosító")
        y -= 16
        if qr_path.exists():
            c.drawImage(str(qr_path), 18 * mm, y - 45 * mm, width=40 * mm, height=40 * mm)
        c.setFont(PDF_FONT, 8)
        y -= 52 * mm
        for k, v in qr_payload.items():
            c.drawString(18 * mm, y, f"{k}: {v}")
            y -= 10
        try:
            qr_path.unlink()
        except Exception:
            pass
        c.showPage()

    c.save()
    return out_pdf

# -----------------------------
# TSV / XLSM output
# -----------------------------

def write_tsv(path: Path, row: Dict[str, str], pdf_path: Optional[Path] = None, warnings: Optional[List[str]] = None) -> Path:
    ensure_dir(path.parent)
    headers = ["__PDF_PATH__", "__PDF_NAME__", "__STATUS__", "__WARNINGS__"] + LETTERS
    data = {k: row.get(k, "") for k in LETTERS}
    data.update({
        "__PDF_PATH__": str(pdf_path or ""),
        "__PDF_NAME__": (pdf_path.name if pdf_path else ""),
        "__STATUS__": "OK",
        "__WARNINGS__": " | ".join(warnings or []),
    })
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=headers, dialect="excel-tab", extrasaction="ignore")
        writer.writeheader()
        writer.writerow(data)
    return path


def write_log(path: Path, pdf_path: Path, person: GeneratedPerson, warnings: Optional[List[str]] = None) -> Path:
    headers = ["Időpont", "PDF fájl", "Állapot", "Név", "Útlevél", "Üzenet"]
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=headers, dialect="excel-tab")
        writer.writeheader()
        writer.writerow({
            "Időpont": datetime.now().strftime("%Y.%m.%d %H:%M:%S"),
            "PDF fájl": pdf_path.name,
            "Állapot": "OK - FIGYELMEZTETÉS" if warnings else "OK",
            "Név": person.full_name,
            "Útlevél": person.passport_no,
            "Üzenet": " | ".join(warnings or []),
        })
    return path


def write_pdf_list(path: Path, pdf_path: Path) -> Path:
    path.write_text(str(pdf_path) + "\n", encoding="utf-8-sig")
    return path


def write_xlsm_copy(workbook_path: Path, out_workbook_path: Path, row_data: Dict[str, str]) -> Optional[Path]:
    if openpyxl is None:
        return None
    if not workbook_path.exists():
        return None
    ensure_dir(out_workbook_path.parent)
    shutil.copy2(workbook_path, out_workbook_path)
    wb = openpyxl.load_workbook(out_workbook_path, keep_vba=True)
    if "Munkavállalók" not in wb.sheetnames:
        raise RuntimeError("Nincs Munkavállalók fül az XLSM-ben.")
    ws = wb["Munkavállalók"]
    # first empty row by C / N / D columns, minimum 2
    target_row = None
    for r in range(2, ws.max_row + 2):
        if not one_line(ws[f"C{r}"].value) and not one_line(ws[f"N{r}"].value):
            target_row = r
            break
    if target_row is None:
        target_row = ws.max_row + 1
    for col, value in row_data.items():
        if col in LETTERS:
            ws[f"{col}{target_row}"] = value
    if "Start" in wb.sheetnames:
        wb["Start"]["C1"] = target_row - 1
    wb.save(out_workbook_path)
    return out_workbook_path

# -----------------------------
# Supabase integration
# -----------------------------

def supabase_headers(profile: Optional[str] = None) -> Dict[str, str]:
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY") or os.environ.get("SUPABASE_ANON_KEY") or ""
    headers = {
        "apikey": key,
        "Authorization": f"Bearer {key}",
    }
    if profile:
        headers["Accept-Profile"] = profile
        headers["Content-Profile"] = profile
    return headers


def storage_upload(local_path: Path, bucket: str, object_path: str, content_type: str = "application/pdf") -> Dict[str, Any]:
    if requests is None:
        return {"ok": False, "reason": "requests_missing"}
    url = os.environ.get("SUPABASE_URL", "").rstrip("/")
    if not url or not os.environ.get("SUPABASE_SERVICE_ROLE_KEY"):
        return {"ok": False, "reason": "supabase_env_missing"}
    endpoint = f"{url}/storage/v1/object/{bucket}/{object_path}"
    h = supabase_headers()
    h.update({"x-upsert": "true", "Content-Type": content_type})
    with local_path.open("rb") as f:
        res = requests.post(endpoint, headers=h, data=f, timeout=120)
    return {
        "ok": res.ok,
        "status_code": res.status_code,
        "text": res.text[:1000],
        "bucket": bucket,
        "object_path": object_path,
        "public_url": f"{url}/storage/v1/object/public/{bucket}/{object_path}",
    }


def register_generated_file(person: GeneratedPerson, local_path: Path, storage_result: Dict[str, Any], file_kind: str) -> Dict[str, Any]:
    if requests is None:
        return {"ok": False, "reason": "requests_missing"}
    url = os.environ.get("SUPABASE_URL", "").rstrip("/")
    if not url or not os.environ.get("SUPABASE_SERVICE_ROLE_KEY"):
        return {"ok": False, "reason": "supabase_env_missing"}
    # Attempt oif_eh schema. This is best-effort because exact generated_files columns may evolve.
    payload = {
        "package_id": person.package_id or None,
        "workflow_case_id": person.workflow_case_id or None,
        "file_kind": file_kind,
        "file_name": local_path.name,
        "storage_bucket": storage_result.get("bucket") or os.environ.get("OIF_EH_GENERATED_BUCKET", "oif-eh-generated"),
        "storage_path": storage_result.get("object_path") or "",
        "storage_url": storage_result.get("public_url") or "",
        "local_path": str(local_path),
        "mime_type": "application/pdf" if local_path.suffix.lower() == ".pdf" else "text/tab-separated-values",
        "file_size_bytes": local_path.stat().st_size if local_path.exists() else None,
        "file_sha256": sha256_file(local_path) if local_path.exists() else None,
        "generation_status": "generated",
        "created_at": datetime.now().isoformat(),
    }
    endpoint = f"{url}/rest/v1/generated_files"
    h = supabase_headers("oif_eh")
    h.update({"Content-Type": "application/json", "Prefer": "return=representation"})
    res = requests.post(endpoint, headers=h, json=payload, timeout=60)
    return {"ok": res.ok, "status_code": res.status_code, "text": res.text[:1000], "payload": payload}

# -----------------------------
# Main generation flow
# -----------------------------

def maybe_run_legacy_extractor(pdf_path: Path, run_dir: Path) -> Dict[str, Any]:
    """Run user's v8 extractor if it is next to this file or in current dir."""
    script_candidates = [
        Path(__file__).resolve().parent / "eh_pdf_extractor_v8_fixed_defaults.py",
        Path.cwd() / "eh_pdf_extractor_v8_fixed_defaults.py",
        Path("/mnt/data/eh_pdf_extractor_v8_fixed_defaults.py"),
    ]
    script = next((p for p in script_candidates if p.exists()), None)
    if not script:
        return {"ok": False, "reason": "extractor_not_found"}
    import subprocess
    input_list = write_pdf_list(run_dir / "pdf_lista.txt", pdf_path)
    out_data = run_dir / "adatok_extractor.tsv"
    out_log = run_dir / "python_log_extractor.tsv"
    cmd = [sys.executable, str(script), "--input-list", str(input_list), "--out-data", str(out_data), "--out-log", str(out_log)]
    proc = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
    return {
        "ok": proc.returncode == 0,
        "returncode": proc.returncode,
        "stdout": proc.stdout[-2000:],
        "stderr": proc.stderr[-2000:],
        "out_data": str(out_data) if out_data.exists() else "",
        "out_log": str(out_log) if out_log.exists() else "",
    }


def generate_package_pdfs(payload: Dict[str, Any], out_dir: Optional[Path] = None, workbook_path: Optional[Path] = None) -> Dict[str, Any]:
    run_id = now_id()
    base_out = out_dir or Path.cwd() / "eh_pdf_generator_runs" / run_id
    ensure_dir(base_out)

    person = person_from_payload(payload)
    unsupported = False
    warnings: List[str] = []
    if person.app_type == "foglalkoztatasi":
        warnings.append("9.6 Foglalkoztatás PDF generálás támogatott. A meglévő legacy EH böngészős kitöltő/extractor továbbra is külön mappinget igényel.")

    person_slug = slugify(person.full_name)
    pdf_name = f"TARTENG_{person_slug}_{person.app_type}_{run_id}.pdf"
    source_pdf = base_out / pdf_name
    generate_filled_source_pdf(person, source_pdf, include_qr_backpage=bool(payload.get("include_qr_backpage", True)))

    row_data = build_xlsm_row(person)
    tsv_path = write_tsv(base_out / "adatok.tsv", row_data, source_pdf, warnings)
    log_path = write_log(base_out / "python_log.tsv", source_pdf, person, warnings)
    pdf_list = write_pdf_list(base_out / "pdf_lista.txt", source_pdf)

    extractor_result = maybe_run_legacy_extractor(source_pdf, base_out)

    generated_workbook = None
    wb_arg = payload.get("workbook_path") or (str(workbook_path) if workbook_path else "")
    if wb_arg:
        try:
            wb_in = Path(str(wb_arg).strip('"'))
            out_wb = base_out / f"ENTERHUNGARY_generated_{person_slug}_{run_id}.xlsm"
            generated_workbook = write_xlsm_copy(wb_in, out_wb, row_data)
        except Exception as exc:
            warnings.append(f"XLSM írás sikertelen: {type(exc).__name__}: {exc}")

    manifest = {
        "ok": not unsupported,
        "status": "unsupported_for_legacy_eh_agent" if unsupported else "generated",
        "run_id": run_id,
        "run_dir": str(base_out),
        "application_type": person.app_type,
        "supported_by_existing_eh_agent": person.app_type in {"vendeg", "nemzeti"},
        "pdf_template_supported": person.app_type in {"vendeg", "nemzeti", "foglalkoztatasi"},
        "package_id": person.package_id,
        "workflow_case_id": person.workflow_case_id,
        "eh_case_id": person.eh_case_id,
        "full_name": person.full_name,
        "source_pdf": str(source_pdf),
        "source_pdf_sha256": sha256_file(source_pdf),
        "tsv_path": str(tsv_path),
        "log_path": str(log_path),
        "pdf_list_path": str(pdf_list),
        "generated_workbook_path": str(generated_workbook) if generated_workbook else "",
        "extractor_result": extractor_result,
        "warnings": warnings,
        "person": asdict(person),
    }

    bucket = os.environ.get("OIF_EH_GENERATED_BUCKET", "oif-eh-generated")
    upload_enabled = bool(payload.get("upload_generated_to_storage", False))
    storage_results: List[Dict[str, Any]] = []
    db_results: List[Dict[str, Any]] = []
    if upload_enabled:
        folder = f"{person.package_id or person.workflow_case_id or run_id}/{run_id}"
        for file_path, kind, ctype in [
            (source_pdf, "filled_source_pdf", "application/pdf"),
            (tsv_path, "xlsm_import_tsv", "text/tab-separated-values"),
            (log_path, "generator_log_tsv", "text/tab-separated-values"),
        ]:
            object_path = f"{folder}/{file_path.name}"
            sr = storage_upload(file_path, bucket, object_path, ctype)
            storage_results.append(sr)
            if sr.get("ok"):
                db_results.append(register_generated_file(person, file_path, sr, kind))
        if generated_workbook:
            sr = storage_upload(generated_workbook, bucket, f"{folder}/{generated_workbook.name}", "application/vnd.ms-excel.sheet.macroEnabled.12")
            storage_results.append(sr)
            if sr.get("ok"):
                db_results.append(register_generated_file(person, generated_workbook, sr, "generated_xlsm"))

    manifest["storage_results"] = storage_results
    manifest["generated_file_db_results"] = db_results
    manifest_path = base_out / "manifest.json"
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    manifest["manifest_path"] = str(manifest_path)
    manifest["generated_file_count"] = 1 + (1 if generated_workbook else 0)
    return manifest

# -----------------------------
# FastAPI
# -----------------------------

if BaseModel is not object:
    class GenerateRequest(BaseModel):
        package_id: Optional[str] = None
        workflow_case_id: Optional[str] = None
        eh_case_id: Optional[str] = None
        application_type: Optional[str] = None
        app_type: Optional[str] = None
        process_kind: Optional[str] = None
        package_type: Optional[str] = None
        python_export_row: Dict[str, Any] = Field(default_factory=dict)
        python_export_rows: List[Dict[str, Any]] = Field(default_factory=list)
        person_folder: Optional[str] = None
        local_person_folder: Optional[str] = None
        project_root: Optional[str] = None
        workbook_path: Optional[str] = None
        include_qr_backpage: bool = True
        upload_generated_to_storage: bool = False
        payload: Dict[str, Any] = Field(default_factory=dict)

if FastAPI is not None:
    app = FastAPI(title="OIF/EH PDF generator front-layer", version="1.0.0")

    @app.get("/health")
    def health() -> Dict[str, Any]:
        return {
            "ok": True,
            "service": "oif_eh_pdf_generator_frontlayer",
            "supported_application_types": ["vendeg", "vendeg_9_7", "nemzeti", "foglalkoztatasi_9_6"],
            "unsupported_yet": [],
            "supabase_configured": bool(os.environ.get("SUPABASE_URL") and os.environ.get("SUPABASE_SERVICE_ROLE_KEY")),
        }

    @app.post("/generate-oif-package-pdfs")
    def api_generate(req: GenerateRequest) -> Dict[str, Any]:  # type: ignore[name-defined]
        d = req.model_dump() if hasattr(req, "model_dump") else req.dict()
        # Merge Retool top-level and python export row while keeping both.
        if d.get("python_export_rows") and not d.get("python_export_row"):
            d["python_export_row"] = d["python_export_rows"][0]
        try:
            return generate_package_pdfs(d, workbook_path=Path(d["workbook_path"]) if d.get("workbook_path") else None)
        except Exception as exc:
            return {
                "ok": False,
                "status": "error",
                "error": f"{type(exc).__name__}: {exc}",
                "traceback": traceback.format_exc(),
            }

# -----------------------------
# CLI
# -----------------------------

def main(argv: Optional[List[str]] = None) -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--payload-json", required=True, help="JSON payload path")
    ap.add_argument("--out-dir", default="", help="Output directory")
    ap.add_argument("--workbook", default="", help="Optional ENTERHUNGARY xlsm path")
    ap.add_argument("--upload-generated-to-storage", action="store_true")
    args = ap.parse_args(argv)

    payload = json.loads(Path(args.payload_json).read_text(encoding="utf-8"))
    if args.workbook:
        payload["workbook_path"] = args.workbook
    if args.upload_generated_to_storage:
        payload["upload_generated_to_storage"] = True
    out_dir = Path(args.out_dir) if args.out_dir else None
    result = generate_package_pdfs(payload, out_dir=out_dir, workbook_path=Path(args.workbook) if args.workbook else None)
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0 if result.get("ok") else 1


if __name__ == "__main__":
    raise SystemExit(main())
