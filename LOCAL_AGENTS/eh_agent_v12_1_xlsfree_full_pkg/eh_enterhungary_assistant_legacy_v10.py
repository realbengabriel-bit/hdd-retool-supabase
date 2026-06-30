# -*- coding: utf-8 -*-
"""
EnterHungary félautomata kitöltő asszisztens - v9 Chrome/CERTFIX

Workflow:
- XLSM-ből beolvassa a Start!C1 szerinti sort a Munkavállalók fülről.
- Eldönti, hogy Nemzeti Kártya vagy Vendégmunkás ügy (DF=36 -> nemzeti, DE=1 -> vendégmunkás).
- Edge-ben / Chromiumban megnyitja az EH új kérelem URL-t.
- Kitölti az 1. oldalt a meglévő Excel mapping fülek alapján.
- Megáll emberi ellenőrzésre.
- Enterre rákattint a Mehet (Rögzít) gombra.
- Megnyitja a betétlap szerkesztő URL-t.
- Kitölti a 2. oldalt.
- Ismét megáll ellenőrzésre.

FONTOS: nem vakon adja be az ügyet; minden Rögzítés előtt kézi ellenőrzési pont van.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import os
import re
import sys
import time
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

try:
    import openpyxl
    from openpyxl.utils.datetime import from_excel
except Exception as exc:  # pragma: no cover
    print("Hiányzik az openpyxl. Telepítés: py -m pip install openpyxl")
    raise

try:
    from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError, Page, BrowserContext
except Exception as exc:  # pragma: no cover
    print("Hiányzik a Playwright. Telepítés:")
    print("  py -m pip install playwright openpyxl")
    print("  py -m playwright install")
    raise

BASE_URL = "https://enterhungary.gov.hu"
NEW_CASE_URLS = {
    # A direkt altípus URL az EH-ban sokszor jó, de vendégmunkásnál login után
    # előfordul, hogy nem a fő kérelem űrlapra dob vissza. Emiatt v4-ben
    # több navigációs fallback is van az open_new_case() alatt.
    "vendeg": BASE_URL + "/eh/cases/new/tartcelharm-c7",
    "nemzeti": BASE_URL + "/eh/cases/new/tartcelharm-c12",
}

NEW_CASE_FALLBACK_URLS = {
    "vendeg": [
        BASE_URL + "/eh/cases/new/tartcelharm-c7",
        BASE_URL + "/eh/cases/new#tartcelharm",
        BASE_URL + "/eh/cases/new/tartcelharm",
    ],
    "nemzeti": [
        BASE_URL + "/eh/cases/new/tartcelharm-c12",
        BASE_URL + "/eh/cases/new#tartcelharm",
        BASE_URL + "/eh/cases/new/tartcelharm",
    ],
}
EDIT_SUFFIX = {
    "vendeg": "tartcelharm-c7",
    "nemzeti": "tartcelharm-c12",
}
FIRST_MAPPING = {
    "vendeg": "VendegmKerelem",
    "nemzeti": "Kerelem",
}
SECOND_MAPPING = {
    "vendeg": "VendegmKerelem2",
    "nemzeti": "Kerelem2",
}
TARGET_SHEET = "Munkavállalók"
START_SHEET = "Start"
LOG_CSV = "eh_asszisztens_log.csv"

DATE_HEADERS = {
    "Születési idő",
    "Útlevél kiállításának dátuma",
    "Útlevél érvényes dátum",
    "Előzetes megállapodás kelte",
    "Meddig kérelmezi tartózkodása engedélyezését?",
    "Nyilvántartásba vétel napja",
    "Beutazás ideje",
}

LITERAL_VALUES = {"igen", "nem", "Igen", "Nem", "foglalkoztató", "minősített kölcsönbeadó", "minősített munkaerő kölcsönző", "Munkáltató"}


def norm(s: Any) -> str:
    """Összehasonlításhoz: kisbetű, ékezet nélkül, whitespace/hyphen normalizálva."""
    if s is None:
        return ""
    s = str(s).strip()
    s = s.replace("–", "-").replace("—", "-")
    s = s.replace("-", " ")
    s = " ".join(s.split())
    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    return s.lower()


def clean_text(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, str):
        return " ".join(v.strip().split())
    return str(v).strip()


def excel_date_to_str(v: Any) -> str:
    if v is None or v == "":
        return ""
    if isinstance(v, dt.datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, dt.date):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, (int, float)) and 20000 < float(v) < 90000:
        try:
            d = from_excel(v)
            if isinstance(d, dt.datetime):
                return d.strftime("%Y-%m-%d")
            if isinstance(d, dt.date):
                return d.strftime("%Y-%m-%d")
        except Exception:
            pass
    s = clean_text(v)
    # 2024.03.11. / 2024-03-11 / 2024/03/11 -> 2024-03-11
    m = re.search(r"(\d{4})[.\-/ ](\d{1,2})[.\-/ ](\d{1,2})", s)
    if m:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    return s


def normalize_phone(v: Any) -> str:
    s = clean_text(v)
    if not s:
        return s
    s = s.replace(" ", "").replace("-", "").replace("/", "")
    if s.startswith("+"):
        return s
    if s.startswith("06"):
        return "+36" + s[2:]
    if s.startswith("36"):
        return "+" + s
    if s.startswith("30") or s.startswith("20") or s.startswith("70"):
        return "+36" + s
    return s


def normalize_dropdown_value(header: str, value: Any, app_type: str) -> str:
    h = norm(header)
    raw = clean_text(value)
    n = norm(raw)

    # Állampolgárság dropdown: az EH nem mindig ugyanazt a PDF/Excel alakot várja.
    if "allampolgarsag" in h:
        if "fulop" in n:
            return "Fülöp Szigeteki"
        if "ukran" in n or "ukraj" in n:
            return "Ukrán"
        if "orosz" in n:
            return "Orosz"
        if "szerb" in n:
            return "Szerb"
        if "mold" in n:
            return "Moldáv"

    # Ország dropdown: itt országnevek kellenek, nem állampolgárságok.
    if "orszag" in h or "celorszaga" in h or "tovabb" in h:
        if "fulop" in n:
            return "Fülöp-Szigetek"
        if "ukran" in n or "ukraj" in n:
            return "Ukrajna"

    if "penznem" in h or h == "penz":
        if n in {"huf", "forint", "ft"} or not raw:
            return "Forint"

    if "neme" in h or h == "nem":
        if n.startswith("no") or "noi" in n:
            return "Nő"
        if n.startswith("ferfi"):
            return "Férfi"

    if "benyujto" in h:
        # EH pontos opciók:
        # - Nemzeti Kártya: foglalkoztató
        # - Vendégmunkás: minősített kölcsönbeadó
        # A felhasználói szóhasználatban ez gyakran "munkáltató" /
        # "minősített munkaerő-kölcsönző" néven szerepel, de a dropdown
        # látható szövegét kell kiválasztani.
        if app_type == "vendeg":
            return "minősített kölcsönbeadó"
        return "foglalkoztató"

    if "minositett" in h and "kolcson" in h:
        return "minősített munkaerő kölcsönző"

    return raw


def normalize_for_field(header: str, value: Any, app_type: str) -> str:
    h = norm(header)
    if any(x in h for x in ["telefon", "telefonszam", "mobilszam"]):
        return normalize_phone(value)
    if header in DATE_HEADERS or "datum" in h or "ideje" in h or "ido" in h or "kelte" in h:
        return excel_date_to_str(value)
    return normalize_dropdown_value(header, value, app_type)


@dataclass
class MappingRow:
    row_no: int
    xpath: str
    kind: str
    col_name: str


class WorkbookData:
    def __init__(self, workbook_path: Path):
        self.path = workbook_path
        self.wb = openpyxl.load_workbook(workbook_path, data_only=True, keep_vba=True)
        if TARGET_SHEET not in self.wb.sheetnames:
            raise RuntimeError(f"Nincs ilyen fül: {TARGET_SHEET}")
        self.people = self.wb[TARGET_SHEET]
        self.headers_by_name: Dict[str, int] = {}
        self.headers_original: Dict[str, str] = {}
        self._read_headers()

    def _read_headers(self) -> None:
        for col in range(1, self.people.max_column + 1):
            val = clean_text(self.people.cell(1, col).value)
            if val:
                self.headers_by_name[norm(val)] = col
                self.headers_original[norm(val)] = val

    def start_row(self) -> int:
        if START_SHEET not in self.wb.sheetnames:
            return 2
        val = self.wb[START_SHEET]["C1"].value
        try:
            return int(val) + 1
        except Exception:
            return 2

    def get_value(self, row_no: int, header_or_literal: str, app_type: str) -> str:
        col_name = clean_text(header_or_literal)
        if not col_name:
            return ""

        key = norm(col_name)
        if key in self.headers_by_name:
            original_header = self.headers_original[key]
            v = self.people.cell(row_no, self.headers_by_name[key]).value
            return normalize_for_field(original_header, v, app_type)

        # Fix literal értékek a mappingből.
        if col_name in LITERAL_VALUES or key in {"igen", "nem"}:
            return col_name

        # Ha nincs ilyen fejléc, ne vigyünk tovább előző értéket.
        return ""

    def get_cell_by_letter(self, row_no: int, col_letter: str) -> str:
        return clean_text(self.people[f"{col_letter}{row_no}"].value)

    def row_display_name(self, row_no: int) -> str:
        for header in ["Teljes Név", "Teljes név", "Név"]:
            v = self.get_value(row_no, header, "")
            if v:
                return v
        return f"Excel sor {row_no}"

    def detect_type(self, row_no: int) -> str:
        de = self.get_cell_by_letter(row_no, "DE")
        df = self.get_cell_by_letter(row_no, "DF")
        if df.strip() == "36":
            return "nemzeti"
        if de.strip() == "1":
            return "vendeg"
        print(f"Nem egyértelmű típus: DE={de!r}, DF={df!r}")
        ans = input("Típus? [n]emzeti / [v]endégmunkás: ").strip().lower()
        return "nemzeti" if ans.startswith("n") else "vendeg"

    def mapping_rows(self, sheet_name: str) -> List[MappingRow]:
        if sheet_name not in self.wb.sheetnames:
            raise RuntimeError(f"Nincs mapping fül: {sheet_name}")
        ws = self.wb[sheet_name]
        rows: List[MappingRow] = []
        for r in range(2, ws.max_row + 1):
            xpath = clean_text(ws.cell(r, 1).value)
            kind = clean_text(ws.cell(r, 2).value)
            col_name = clean_text(ws.cell(r, 3).value)
            if not xpath and not kind and not col_name:
                continue
            if not xpath:
                continue
            rows.append(MappingRow(r, xpath, kind, col_name))
        return rows


def append_log(base_dir: Path, row: List[Any]) -> None:
    log_path = base_dir / LOG_CSV
    exists = log_path.exists()
    with open(log_path, "a", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        if not exists:
            w.writerow(["timestamp", "excel_row", "name", "app_type", "stage", "status", "detail", "url"])
        w.writerow(row)


class EHAssistant:
    def __init__(self, page: Page, data: WorkbookData, row_no: int, app_type: str, base_dir: Path):
        self.page = page
        self.data = data
        self.row_no = row_no
        self.app_type = app_type
        self.base_dir = base_dir
        self.name = data.row_display_name(row_no)

    def log(self, stage: str, status: str, detail: str = "") -> None:
        ts = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        append_log(self.base_dir, [ts, self.row_no, self.name, self.app_type, stage, status, detail, self.page.url])
        print(f"[{status}] {stage}: {detail}")

    def is_login_page(self) -> bool:
        """Az EH néha nem login szót tartalmazó URL-re dob, ezért nem csak URL alapján nézzük."""
        url = self.page.url.lower()
        if "login" in url or "belepes" in url or "auth" in url:
            return True
        try:
            if self.page.locator("input[type='password']").count() > 0:
                return True
        except Exception:
            pass
        try:
            body = (self.page.locator("body").inner_text(timeout=1500) or "").lower()
            if "jelszó" in body or "jelszo" in body or "bejelentkez" in body:
                return True
        except Exception:
            pass
        return False

    def wait_for_login_if_needed(self, target_url: Optional[str] = None) -> None:
        time.sleep(1)
        if self.is_login_page():
            print("\nÚgy tűnik, bejelentkezési oldalon vagy.")
            print("Lépj be az EH-ban. Ha belépés után nem oda dob vissza, ahová kell, semmi gond: Enter után visszanavigálok a cél URL-re.")
            input("Bejelentkezés után itt nyomj Entert...")
            if target_url:
                print(f"Visszanavigálás céloldalra: {target_url}")
                self.safe_goto(target_url)
                time.sleep(1)

    def safe_wait_dom(self) -> None:
        try:
            self.page.wait_for_load_state("domcontentloaded", timeout=20000)
        except Exception:
            pass

    def disable_leave_warning(self) -> None:
        """Az EH szerkesztő oldal elhagyásakor felugró confirm néha bezavar a Playwrightba."""
        try:
            self.page.evaluate(
                """
                () => {
                    window.onbeforeunload = null;
                    window.onunload = null;
                    document.querySelectorAll('[data-confirmleave]').forEach(e => e.removeAttribute('data-confirmleave'));
                }
                """
            )
        except Exception:
            pass

    def safe_goto(self, url: str) -> None:
        self.disable_leave_warning()
        try:
            self.page.goto(url, wait_until="domcontentloaded", timeout=60000)
        except Exception as exc:
            # Ha közben EH confirm dialog vagy átirányítás bezavar, még egyszer próbáljuk.
            print(f"Navigáció első próbára nem sikerült: {exc}")
            time.sleep(1)
            self.disable_leave_warning()
            self.page.goto(url, wait_until="domcontentloaded", timeout=60000)
        self.disable_leave_warning()

    def has_any_field(self, names: List[str], timeout_ms: int = 3000) -> bool:
        for name in names:
            try:
                self.page.locator(f"[name='{name}']").first.wait_for(state="attached", timeout=timeout_ms)
                return True
            except Exception:
                pass
        return False

    def try_click_new_case_type_link(self) -> bool:
        """Új kérelem választóoldalon megpróbálja kiválasztani a megfelelő típust."""
        href_part = "tartcelharm-c7" if self.app_type == "vendeg" else "tartcelharm-c12"
        text_part = "Vendégmunkás" if self.app_type == "vendeg" else "Nemzeti Kártya"
        candidates = [
            f"a[href*='{href_part}']",
            f"button:has-text('{text_part}')",
            f"a:has-text('{text_part}')",
            f"input[value*='{text_part}']",
        ]
        for sel in candidates:
            try:
                loc = self.page.locator(sel).first
                if loc.count() > 0:
                    loc.click(timeout=5000)
                    self.safe_wait_dom()
                    self.disable_leave_warning()
                    return True
            except Exception:
                pass
        return False

    def ensure_form_ready(self, stage: str, expected_names: List[str], target_url: Optional[str] = None) -> None:
        """Megvárja, hogy tényleg a kitöltő űrlapon legyünk. Vendégmunkásnál login után néha rossz oldalra dob vissza."""
        for attempt in range(1, 4):
            for name in expected_names:
                try:
                    self.page.locator(f"[name='{name}']").first.wait_for(state="attached", timeout=5000)
                    return
                except Exception:
                    pass
            if target_url and attempt < 3:
                print(f"Nem látom még a(z) {stage} űrlapot. Újrapróbálom: {target_url}")
                self.safe_goto(target_url)
                self.wait_for_login_if_needed(target_url)
                time.sleep(1)

        print("\nNem látom a várt EH űrlapmezőket.")
        print(f"Aktuális URL: {self.page.url}")
        try:
            print(f"Oldalcím: {self.page.title()}")
        except Exception:
            pass
        input("Navigálj kézzel a megfelelő EH űrlapra, majd itt nyomj Entert...")

    def open_new_case(self) -> None:
        """Új ügy nyitása robusztusan.

        Vendégmunkásnál az EH login után néha nem közvetlenül a kitöltő űrlapra
        visz vissza. Ilyenkor a v4 végigpróbálja a releváns új ügy URL-eket,
        és csak akkor kér kézi beavatkozást, ha tényleg nem látszanak az 1. oldal
        fő mezői.
        """
        expected = ["kerelmezocsaladnev", "kerelmezoutonev", "kerelmezoap"]
        candidates = NEW_CASE_FALLBACK_URLS.get(self.app_type, [NEW_CASE_URLS[self.app_type]])
        last_url = candidates[0]

        for idx, url in enumerate(candidates, start=1):
            last_url = url
            print(f"Megnyitás / próba {idx}: {url}")
            self.safe_goto(url)
            self.wait_for_login_if_needed(url)
            self.safe_wait_dom()
            self.disable_leave_warning()
            if self.has_any_field(expected, timeout_ms=8000):
                self.log("open_new_case", "OK", self.page.url)
                return

            # Ha az EH új ügy választóoldalon vagyunk, próbáljuk kattintani a konkrét altípus linkjét.
            if self.try_click_new_case_type_link():
                self.wait_for_login_if_needed(self.page.url)
                self.safe_wait_dom()
                self.disable_leave_warning()
                if self.has_any_field(expected, timeout_ms=8000):
                    self.log("open_new_case", "OK", self.page.url)
                    return

        self.ensure_form_ready("1. oldal", expected, last_url)
        self.log("open_new_case", "OK", self.page.url)

    def value_for_mapping(self, m: MappingRow) -> str:
        val = self.data.get_value(self.row_no, m.col_name, self.app_type)
        default = self.default_for_mapping(m)
        if default is not None:
            return default
        return val

    def default_for_mapping(self, m: MappingRow) -> Optional[str]:
        sh = ""  # sheet-specific default most often applied by name later
        c = norm(m.col_name)
        if "szallashelyen tartozkodas jogcime" in c or "tartozkodas jogcime" in c:
            return "Egyéb"
        if c == norm("Éspedig"):
            return "befogadó nyilatkozat, munkáltató bérli"
        if "rendelkezik e egeszsegbiztositassal" in c or "egeszsegbiztositas" in c:
            return "Foglalkoztatási Jogviszony Alapján"
        if "penznem" in c:
            return "Forint"
        return None

    def should_skip(self, sheet_name: str, m: MappingRow) -> bool:
        sh = norm(sheet_name)
        c = norm(m.col_name)
        x = m.xpath.lower()
        # 242 rubrika maradjon üres, mert KH kérdés default Nem.
        if sh == "kerelem2" and "2023 evi xc torveny 242" in c:
            return True
        # Postai kézbesítés címe / meghatalmazott kapcsolattartási címe:
        # korábban vendégmunkásnál kihagytuk, v7-től kérés szerint pipáljuk.
        # MO előtti tartózkodási hely közterület neve nem kötelező, maradjon üres.
        if "moelotttarthelykozteruletneve" in x:
            return True
        if "magyarorszagra erkezest megelozo" in c and "kozterulet neve" in c:
            return True

        # Vendégmunkás 2. oldal: a mapping 47. sora Állampolgárság néven fut,
        # de az XPath valójában az anyanyelv dropdownra mutat. Ezt defaultból töltjük: Angol.
        # Így nincs felesleges figyelmeztetés és nem vár a rossz opcióra.
        if sh == "vendegmkerelem2" and "allampolgarsag" in c:
            return True

        return False

    def fill_mapping_sheet(self, sheet_name: str) -> None:
        rows = self.data.mapping_rows(sheet_name)
        self.log("fill_mapping", "START", sheet_name)
        for m in rows:
            if self.should_skip(sheet_name, m):
                self.log(sheet_name, "SKIP", f"{m.row_no} {m.col_name}")
                continue
            value = self.value_for_mapping(m)
            kind = self.normalize_kind(m.kind, m.xpath)
            try:
                self.fill_element(m.xpath, kind, value, sheet_name, m.row_no)
            except Exception as exc:
                self.log(sheet_name, "WARN", f"{m.row_no} {kind} {m.col_name}: {exc}")
        self.apply_page_defaults(sheet_name)
        self.log("fill_mapping", "DONE", sheet_name)

    def normalize_kind(self, kind: str, xpath: str) -> str:
        k = clean_text(kind)
        if norm(k) == "input" and "select" in xpath.lower():
            return "Combobox"
        return k

    def fill_element(self, xpath: str, kind: str, value: str, sheet_name: str, row_no: int) -> None:
        loc = self.page.locator(f"xpath={xpath}").first
        k = norm(kind)
        if not value and k != "radio":
            return

        # v5: ha egy mapping XPath feltételes / nem létező mezőre mutat, ne várjunk 30 mp-et.
        # Ez a vendégmunkás első oldalon gyorsítja a nem megjelenő blokkokat.
        try:
            loc.wait_for(state="attached", timeout=900)
        except Exception as exc:
            raise RuntimeError(f"XPath nem található 0.9 mp alatt: {xpath}") from exc

        if k == "radio":
            loc.click(timeout=900)
            return
        if k == "combobox":
            self.select_by_locator(loc, value)
            return
        # input/text
        try:
            loc.click(timeout=700)
        except Exception:
            pass
        try:
            loc.fill(str(value), timeout=1500)
        except Exception:
            # JS fallback
            loc.evaluate("(e, v) => { e.value = v; e.dispatchEvent(new Event('input', {bubbles:true})); e.dispatchEvent(new Event('change', {bubbles:true})); e.blur(); }", str(value), timeout=1500)

    def select_by_locator(self, loc, value: str) -> None:
        if value is None or value == "":
            return
        v = clean_text(value)
        try:
            loc.wait_for(state="attached", timeout=900)
        except Exception as exc:
            raise RuntimeError("Dropdown mező nem található 0.9 mp alatt") from exc
        # Első próbálkozás: label szerint.
        try:
            loc.select_option(label=v, timeout=1200)
            return
        except Exception:
            pass
        # Második: value szerint.
        try:
            loc.select_option(value=v, timeout=800)
            return
        except Exception:
            pass
        # Harmadik: normalizált option-text/value keresés JS-sel.
        wanted = norm(v)
        ok = loc.evaluate(
            """
            (select, wanted) => {
                function norm(s) {
                    return (s || '')
                      .normalize('NFD').replace(/[\u0300-\u036f]/g, '')
                      .replace(/[-–—]/g, ' ')
                      .replace(/\s+/g, ' ')
                      .trim().toLowerCase();
                }
                for (const opt of select.options) {
                    if (norm(opt.textContent) === wanted || norm(opt.value) === wanted) {
                        select.value = opt.value;
                        select.dispatchEvent(new Event('change', {bubbles:true}));
                        select.dispatchEvent(new Event('input', {bubbles:true}));
                        return true;
                    }
                }
                // Fülöp-szigeteki / Fülöp Szigeteki jellegű eltérésekre részleges fallback
                for (const opt of select.options) {
                    if (wanted && (norm(opt.textContent).includes(wanted) || wanted.includes(norm(opt.textContent)))) {
                        select.value = opt.value;
                        select.dispatchEvent(new Event('change', {bubbles:true}));
                        select.dispatchEvent(new Event('input', {bubbles:true}));
                        return true;
                    }
                }
                return false;
            }
            """,
            wanted,
            timeout=1500,
        )
        if not ok:
            raise RuntimeError(f"Dropdown opció nem található: {v}")

    def fill_by_name(self, name: str, value: str, allow_empty: bool = False) -> None:
        if not value and not allow_empty:
            return
        loc = self.page.locator(f"[name='{name}']").first
        try:
            loc.wait_for(state="attached", timeout=2500)
            try:
                loc.fill(value, timeout=1500)
            except Exception:
                loc.evaluate("(e, v) => { e.value = v; e.dispatchEvent(new Event('input', {bubbles:true})); e.dispatchEvent(new Event('change', {bubbles:true})); e.blur(); }", value)
        except Exception as exc:
            self.log("default_input", "WARN", f"{name}={value}: {exc}")

    def select_by_name(self, name: str, value: str) -> None:
        try:
            loc = self.page.locator(f"select[name='{name}']").first
            self.select_by_locator(loc, value)
        except Exception as exc:
            self.log("default_select", "WARN", f"{name}={value}: {exc}")

    def has_name(self, name: str) -> bool:
        try:
            return self.page.locator(f"[name='{name}']").count() > 0
        except Exception:
            return False

    def radio_by_name(self, name: str, value: str, warn: bool = True) -> None:
        try:
            loc = self.page.locator(f"input[type='radio'][name='{name}'][value='{value}']").first
            if loc.count() == 0:
                if warn:
                    self.log("default_radio", "SKIP", f"{name}={value}: nincs ilyen mező ezen az oldalon")
                return
            loc.click(timeout=900)
        except Exception as exc:
            if warn:
                self.log("default_radio", "WARN", f"{name}={value}: {exc}")

    def apply_page_defaults(self, sheet_name: str) -> None:
        sh = norm(sheet_name)
        if sh in {"kerelem", "vendegmkerelem"}:
            self.radio_by_name("nemfizetek", "nem")
            self.radio_by_name("tartenghosszab", "nem")
            # Határmenti mező csak ott van töltve, ahol tényleg létezik. Vendégmunkásnál ne lassítson warninggal.
            self.radio_by_name("hatarmentiingazo", "nem", warn=False)
            # A vissza-/továbbutazás feltételei blokkban ezek néha rossz XPathon vannak a régi mappingben.
            self.radio_by_name("utlevel", "igen", warn=False)
            self.radio_by_name("vizum", "igen", warn=False)
            self.radio_by_name("jegy", "nem", warn=False)
            self.select_by_name("tartjogcim", "Egyéb")
            self.fill_by_name("tartjogcimegyeb", "befogadó nyilatkozat, munkáltató bérli")
            self.select_by_name("ebmagyartart", "Foglalkoztatási Jogviszony Alapján")
            self.fill_by_name("moelotttarthelykozteruletneve", "", allow_empty=True)

            # Okmány átvétele: postai kézbesítés + meghatalmazott kapcsolattartási címe.
            # Az EH HTML-ben az érték neve elgépelve szerepel: cim_meghatalamzott.
            self.radio_by_name("atvetel", "atvetel_posta", warn=False)
            self.radio_by_name("atvetel_cim", "cim_meghatalamzott", warn=False)

            if sh == "vendegmkerelem":
                self.radio_by_name("penz", "igen")
                self.fill_by_name("osszeg", "100000")
                self.select_by_name("osszegpenznem", "Forint")

        if sh == "kerelem2":
            self.radio_by_name("tudmagyarul", "nem")
            # Nemzeti kártya: anyanyelv 99%-ban állampolgárság alapján.
            citizenship = self.data.get_value(self.row_no, "Állampolgárság", self.app_type)
            mother_tongue = "Ukrán" if "ukr" in norm(citizenship) else citizenship
            if mother_tongue:
                self.select_by_name("anyanyelv", mother_tongue)
            # A 242. § KH kérdés: Nem, rubrika üresen.
            for possible_name in ["mentessegkhiv", "kormanyhivatalnemmukodik", "szakhatmentes"]:
                self.radio_by_name(possible_name, "nem")
            for possible_name in ["mentessegkhivpont", "szakhatmentespont", "khivpont"]:
                self.fill_by_name(possible_name, "", allow_empty=True)

        if sh == "vendegmkerelem2":
            self.radio_by_name("tudmagyarul", "nem")
            self.select_by_name("anyanyelv", "Angol")

            # Kinti vízum / útlevél átvétel városa vendégmunkásnál fix.
            self.fill_by_name("vizumatvetelvaros", "malina")

            self.radio_by_name("btatv34", "igen")
            time.sleep(0.3)
            self.fill_by_name("btatv34nap", "2024-03-11")
            self.fill_by_name("btatv34szam", "BP/0702/00010-4/2024")

    def confirm_and_submit(self, stage_label: str) -> bool:
        print("\n" + "=" * 70)
        print(f"ELLENŐRZÉSI PONT: {stage_label}")
        print("Nézd át az EnterHungary oldalon. Enter = Mehet (Rögzít), s = kihagyja a rögzítést, q = kilép.")
        ans = input("> ").strip().lower()
        if ans == "q":
            return False
        if ans == "s":
            self.log(stage_label, "SKIP_SUBMIT", "Felhasználó kihagyta")
            return True
        self.click_submit()
        self.log(stage_label, "SUBMITTED", "Mehet (Rögzít)")
        return True

    def click_submit(self) -> None:
        candidates = [
            "button:has-text('Mehet')",
            "input[type='submit'][value*='Mehet']",
            "button.btn-success",
            "//button[contains(normalize-space(.), 'Mehet')]",
        ]
        last_err = None
        for sel in candidates:
            try:
                loc = self.page.locator(sel).first if not sel.startswith("//") else self.page.locator(f"xpath={sel}").first
                loc.click(timeout=5000)
                self.page.wait_for_load_state("domcontentloaded", timeout=20000)
                time.sleep(1)
                return
            except Exception as exc:
                last_err = exc
        raise RuntimeError(f"Nem találom a Mehet/Rögzít gombot: {last_err}")

    def extract_case_id(self) -> Optional[str]:
        m = re.search(r"/cases/(?:edit|put)/(\d+)", self.page.url)
        if m:
            return m.group(1)
        # linkekből is próbáljuk
        try:
            hrefs = self.page.locator("a[href*='/eh/cases/edit/']").evaluate_all("els => els.map(a => a.href)")
            for h in hrefs:
                m = re.search(r"/cases/edit/(\d+)", h)
                if m:
                    return m.group(1)
        except Exception:
            pass
        return None

    def open_second_page(self) -> bool:
        case_id = self.extract_case_id()
        if not case_id:
            print("Nem tudtam automatikusan kiolvasni az EH ügyazonosítót az URL-ből.")
            case_id = input("Írd be az EH ügyazonosítót, vagy Enter a megszakításhoz: ").strip()
            if not case_id:
                return False
        suffix = EDIT_SUFFIX[self.app_type]
        url = f"{BASE_URL}/eh/cases/edit/{case_id}/{suffix}"
        print(f"Betétlap megnyitása: {url}")
        self.safe_goto(url)
        self.wait_for_login_if_needed(url)
        if self.app_type == "vendeg":
            self.ensure_form_ready("2. oldal / vendégmunkás betétlap", ["anyanyelv", "btatv34", "btatv34nap", "btatv34szam"], url)
        else:
            self.ensure_form_ready("2. oldal / nemzeti betétlap", ["anyanyelv", "tudmagyarul"], url)
        self.log("open_second_page", "OK", url)
        return True

    def run(self) -> None:
        self.open_new_case()
        self.fill_mapping_sheet(FIRST_MAPPING[self.app_type])
        if not self.confirm_and_submit("1. oldal / fő kérelem"):
            return
        if not self.open_second_page():
            return
        self.fill_mapping_sheet(SECOND_MAPPING[self.app_type])
        if not self.confirm_and_submit("2. oldal / betétlap"):
            return
        self.log("workflow", "DONE", "Ügy kitöltési workflow befejezve")
        print("\nKész. Ellenőrizd az EH-ban a végső állapotot / dokumentumfeltöltést / beadást.")


def find_default_workbook(script_dir: Path) -> Optional[Path]:
    candidates = list(script_dir.glob("*.xlsm"))
    if not candidates:
        return None
    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0]


def safe_accept_dialog(dialog) -> None:
    try:
        print(f"EH felugró ablak kezelve: {dialog.message}")
        dialog.accept()
    except Exception:
        pass


def main() -> int:
    parser = argparse.ArgumentParser(description="EnterHungary félautomata kitöltő asszisztens")
    parser.add_argument("--workbook", "-w", help="XLSM elérési út. Ha nincs megadva, a script mappájában lévő legfrissebb XLSM-et használja.")
    parser.add_argument("--row", "-r", type=int, help="Excel fizikai sor száma a Munkavállalók fülön. Ha nincs megadva: Start!C1 + 1.")
    parser.add_argument("--type", "-t", choices=["nemzeti", "vendeg"], help="Kérelemtípus felülírása.")
    parser.add_argument("--profile", default=".eh_edge_profile", help="Playwright böngészőprofil mappa, ebben marad meg az EH login session.")
    parser.add_argument("--browser", choices=["chrome", "edge", "chromium", "auto"], default="edge", help="Melyik böngészőt használja. Alapértelmezett: edge.")
    parser.add_argument("--connect-cdp", action="store_true", help="Már futó Edge/Chrome remote-debugging böngészőhöz kapcsolódik. Akkor használd, ha Playwright által indított böngészőn nem kattintható a belépés gomb.")
    parser.add_argument("--cdp-url", default="http://127.0.0.1:9222", help="Remote debugging CDP URL. Alap: http://127.0.0.1:9222")
    args = parser.parse_args()

    script_dir = Path(__file__).resolve().parent
    wb_path = Path(args.workbook).resolve() if args.workbook else find_default_workbook(script_dir)
    if not wb_path or not wb_path.exists():
        print("Nem találom az XLSM-et. Add meg így:")
        print("  py eh_enterhungary_assistant.py --workbook ENTERHUNGARY_20260519_fin.xlsm")
        return 2

    data = WorkbookData(wb_path)
    row_no = args.row or data.start_row()
    app_type = args.type or data.detect_type(row_no)
    name = data.row_display_name(row_no)

    print("\nEH ASSZISZTENS")
    print(f"Workbook: {wb_path}")
    print(f"Excel sor: {row_no}")
    print(f"Munkavállaló: {name}")
    print(f"Típus: {'Nemzeti Kártya' if app_type == 'nemzeti' else 'Vendégmunkás'}")
    print("Excelből indítva a gomb már mentette a munkafüzetet. Indítás...")
    if args.connect_cdp:
        print(f"CDP mód: már futó böngészőhöz kapcsolódok: {args.cdp_url}")

    profile_dir = (script_dir / args.profile).resolve()
    launch_args = [
        "--ignore-certificate-errors",
        "--allow-running-insecure-content",
        "--disable-features=HttpsFirstBalancedModeAutoEnable,HttpsUpgrades",
    ]

    def launch_with_channel(playwright_obj, channel_name: Optional[str], label: str) -> BrowserContext:
        kwargs = dict(
            user_data_dir=str(profile_dir),
            headless=False,
            viewport={"width": 1400, "height": 950},
            accept_downloads=True,
            ignore_https_errors=True,
            args=launch_args,
            chromium_sandbox=True,
        )
        if channel_name:
            kwargs["channel"] = channel_name
        print(f"Böngésző indítása: {label} | profil: {profile_dir}")
        return playwright_obj.chromium.launch_persistent_context(**kwargs)

    with sync_playwright() as p:
        context: BrowserContext
        browser = None
        close_context_at_end = True

        if args.connect_cdp:
            # CDP mód: a felhasználó által/StartEdge-del indított, normál Edge-hez csatlakozunk.
            # Ez akkor kell, ha a Playwright által indított izolált profilban az EH login gombja nem kattintható.
            try:
                browser = p.chromium.connect_over_cdp(args.cdp_url, timeout=15000)
                if browser.contexts:
                    context = browser.contexts[0]
                else:
                    context = browser.new_context(ignore_https_errors=True, accept_downloads=True)
                close_context_at_end = False
                print("CDP kapcsolat OK. A meglévő Edge ablakot használom, nem zárom be a végén.")
            except Exception as exc:
                print(f"Nem sikerült csatlakozni a futó Edge-hez CDP-n ({exc}).")
                print("Indítsd el előbb a StartEdge_RemoteDebug_EH.vbs fájlt, majd próbáld újra a CDP gombbal.")
                return 3
        else:
            try:
                if args.browser == "chrome":
                    context = launch_with_channel(p, "chrome", "Google Chrome")
                elif args.browser == "edge":
                    context = launch_with_channel(p, "msedge", "Microsoft Edge")
                elif args.browser == "chromium":
                    context = launch_with_channel(p, None, "Playwright Chromium")
                else:
                    try:
                        context = launch_with_channel(p, "msedge", "Microsoft Edge")
                    except Exception as edge_exc:
                        print(f"Edge indítás nem sikerült ({edge_exc}). Próbálom Chrome-mal.")
                        context = launch_with_channel(p, "chrome", "Google Chrome")
            except Exception as exc:
                print(f"A választott böngésző indítása nem sikerült ({exc}). Fallback Playwright Chromiumra.")
                context = launch_with_channel(p, None, "Playwright Chromium")

        page = context.pages[0] if context.pages else context.new_page()
        # EH oldalelhagyási confirm / alert biztonságos kezelése.
        page.on("dialog", lambda dialog: safe_accept_dialog(dialog))
        assistant = EHAssistant(page, data, row_no, app_type, wb_path.parent)
        try:
            assistant.run()
        finally:
            input("Enter a script befejezéséhez...")
            if close_context_at_end:
                try:
                    context.close()
                except Exception as exc:
                    print(f"A böngésző már bezáródott / kapcsolat megszakadt, ezt figyelmen kívül hagyom: {exc}")
            else:
                print("CDP módban nem zárom be a böngészőt.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
