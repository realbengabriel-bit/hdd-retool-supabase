# -*- coding: utf-8 -*-
r"""
EH melléklet asszisztens v9 PROJECTROOT + TARTENG FIX

Mit csinál:
- Lokális mappából vagy ZIP-ből felismeri a dokumentumokat.
- EnterHungary Fájlmellékletek fülön a live HTML alapján keresi a doktip sorokat.
- Ha a dokumentum már a tárhely selectben van, kiválasztja.
- Ha nincs, a sorhoz tartozó Feltölt linkre megy, feltölti, visszatér, kiválasztja.
- Végén csak külön jóváhagyással submitolja a melléklet formot.
- CERTFIX: az EH tanúsítvány/dátum hibája esetén a Playwright kontextus ignore_https_errors=True beállítással indul.
- PHOTOFALLBACK: ha van fotó PNG/JPG, akkor az megy az arckép/fotó sorba, a kérelem PDF pedig Egyéb 1-be. Ha nincs fotó kép, akkor marad a v6 logika: kérelem PDF megy az arckép/fotó sorba.
- FIXDOCS: Egyéb 4 és Egyéb 5 továbbra is fix dokumentum.
- PROJECTROOT FIX: --project-root/--root használatnál nem dob AttributeError hibát.
- TARTENG FIX: a tart.eng / tart eng fájlnevet kérelem PDF-ként ismeri fel.

Használat:
  py eh_melleklet_assistant_v9_projectroot_tarteng_fix.py --case-id 19020197 --person-folder "C:\Drive\SEWS_10fo\ARGUILLA MICHELLE"
  py eh_melleklet_assistant_v9_projectroot_tarteng_fix.py --person-folder "G:\...\SEWS 10 fő\ARGUILLA MICHELLE"
  py eh_melleklet_assistant_v9_projectroot_tarteng_fix.py --excel ENTERHUNGARY.xlsm --project-root "G:\Megosztott meghajtók\Hatóság\2026\Projektek (jelöltek teljes anyagai)\SEWS 10 fő"
  py eh_melleklet_assistant_v9_projectroot_tarteng_fix.py --person-name "ARGUILLA MICHELLE" --project-root "G:\...\SEWS 10 fő"
"""
from __future__ import annotations

import argparse
import csv
import os
import re
import shutil
import tempfile
import time
import unicodedata
import zipfile
from dataclasses import dataclass
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Dict, List, Optional, Tuple

try:
    from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
except Exception:
    print("HIÁNYZÓ FÜGGŐSÉG: playwright nincs telepítve.")
    print("Futtasd: py -m pip install playwright && py -m playwright install chromium")
    raise

BASE_URL = "https://enterhungary.gov.hu"
SUPPORTED_EXT = {".pdf", ".png", ".jpg", ".jpeg"}
MIN_SCORE_AUTO = 70

# Fix, mindenkire feltöltendő dokumentumok.
# Ezeket a script a személymappa szülői között megkeresi egészen a
# "Projektek (jelöltek teljes anyagai)" mappáig, plusz a script mappájában is.
FIXED_DOCS: Dict[str, Dict[str, object]] = {
    "egyebszukseges4": {
        "canonical": "fix PD meghatalmazás OIF",
        "target_name": "PD meghatalmazás OIF",
        "filenames": [
            "PD meghatalmazás OIF.pdf",
            "PD meghatalmazas OIF.pdf",
            "meghatalmazás OIF.pdf",
            "meghatalmazas OIF.pdf",
        ],
        "priority": 94,
    },
    "egyebszukseges5": {
        "canonical": "fix Posta kérelem_JZ",
        "target_name": "Posta kérelem_JZ",
        "filenames": [
            "Posta kérelem_JZ.pdf",
            "Posta kerelem_JZ.pdf",
            "Posta kérelem JZ.pdf",
            "Posta kerelem JZ.pdf",
        ],
        "priority": 95,
    },
}

# A doktip-eket a beküldött vendégmunkás/nemzeti Mellékletek HTML alapján pontosítottuk.
# A hiányzó/ritkább doktip-ek nem zavarnak: csak akkor töltünk, ha a live EH oldalon tényleg van ilyen sor.
DOC_RULES: Dict[str, Dict[str, object]] = {
    "utlevel": {
        "canonical": "érvényes útlevél teljes másolata",
        "aliases": ["utlevel", "útlevél", "passport", "utlevelteljes", "utlevel teljes", "teljes utlevel", "utiokmany", "úti okmány"],
        "target_name": "útlevél",
        "priority": 10,
    },
    "arckep": {
        # Alaplogika: ha nincs külön fotó kép, ide megy a kérelem PDF.
        # Ha van PNG/JPG fotó, akkor később policy alapján a PNG/JPG megy ide,
        # és a kérelem PDF átmegy az egyebszukseges1 sorba.
        "canonical": "arckép/fotó vagy kérelem PDF",
        "aliases": ["kerelem", "kérelem", "tarteng", "tart eng", "tart.eng", "tart engedely", "tart engedély", "tartozkodasi engedely", "tartózkodási engedély", "kerelmi", "kérelmi", "foto", "fotó", "photo", "arckep", "arckép", "kep", "kép"],
        "target_name": "arckép/fotó vagy kérelem",
        "priority": 20,
    },
    "meghatalmazas": {
        "canonical": "meghatalmazás",
        "aliases": ["meghatalmazas", "meghatalmazás", "meghat", "meghatalm", "mehtalmazas", "meghatálmatzás", "meghatalmaz", "meghatalmazéás", "meghatalmazs"],
        "target_name": "meghatalmazás",
        "priority": 30,
    },
    "fogljogviszonyelozetesmegallap": {
        "canonical": "előzetes megállapodás",
        "aliases": ["elozetes", "előzetes", "megallapodas", "megállapodás", "elozetes megallapodas", "elomegallapodas", "előzetes megall", "elözetes"],
        "target_name": "előzetes megállapodás",
        "priority": 40,
    },
    "szallashigazlap": {
        "canonical": "szálláshely meglétét igazoló okirat / szálláshely bejelentés",
        "aliases": ["szallashely", "szálláshely", "szallas bejelentes", "szállás bejelentés", "szallasbejelentes", "bejelentes", "szallas", "szállás", "szallash", "szállásh"],
        "target_name": "szálláshely bejelentés",
        "priority": 50,
    },
    "lakasberletszerz": {
        "canonical": "lakásbérleti szerződés",
        "aliases": ["berleti", "bérleti", "berlet", "bérlet", "lakasberlet", "lakásbérlet", "szerzodes", "szerződés", "albérlet", "alberlet"],
        "target_name": "bérleti szerződés",
        "priority": 55,
    },
    "lakastulajdlap": {
        "canonical": "tulajdoni lap",
        "aliases": ["tulajdoni", "tullap", "tul lap", "tulajdonilap", "tulajdni", "tulajdoni lap"],
        "target_name": "tulajdoni lap",
        "priority": 56,
    },
    "egeszsegellat01": {
        "canonical": "TAJ / társadalombiztosítási kártya",
        "aliases": ["taj", "egeszseg", "egészség", "tb", "biztositas", "biztosítás", "tarsadalombiztositas", "társadalombiztosítás"],
        "target_name": "TAJ",
        "priority": 60,
    },
    "egbiztkotv": {
        "canonical": "egészségbiztosítási kötvény",
        "aliases": ["kotveny", "kötvény", "biztositasi kotveny", "biztosítási kötvény", "egbizt"],
        "target_name": "egészségbiztosítási kötvény",
        "priority": 61,
    },
    "munkjovig": {
        "canonical": "munkáltatói / 6 havi jövedelemigazolás",
        "aliases": ["6havi", "6 havi", "hat havi", "jovedelem", "jövedelem", "munkaltatoi", "munkáltatói", "kereseti", "jovedig"],
        "target_name": "6 havi jövedelemigazolás",
        "priority": 70,
    },
    "adohatjovedigaz": {
        "canonical": "M30 / adóhatósági jövedelemigazolás",
        "aliases": ["m30", "ado", "adó", "adohatosagi", "adóhatósági", "nav"],
        "target_name": "M30",
        "priority": 80,
    },
    "penzintigaz": {
        "canonical": "pénzintézeti igazolás / bankszámla",
        "aliases": ["bankszamla", "bankszámla", "bank", "penzintezeti", "pénzintézeti", "kivonat", "szamlaegyenleg", "számlaegyenleg"],
        "target_name": "bankszámla igazolás",
        "priority": 82,
    },
    "utazasfedezes": {
        "canonical": "utazási költségek fedezete",
        "aliases": ["utazasfedezet", "utazás fedezet", "utazasi", "utazási", "menetjegy", "jegy"],
        "target_name": "utazás fedezete",
        "priority": 83,
    },
    "szakkpepzesigazolas": {
        "canonical": "szakképzettséget igazoló okirat",
        "aliases": ["szakkepzettseg", "szakképzettség", "vegzettseg", "végzettség", "diploma", "oklevel", "oklevél"],
        "target_name": "szakképzettség igazolás",
        "priority": 84,
    },
    "kiskorutarthozzajarulas": {
        "canonical": "kiskorú hozzájárulás",
        "aliases": ["kiskoru", "kiskorú", "hozzajarulas", "hozzájárulás", "szulo", "szülő"],
        "target_name": "kiskorú hozzájárulás",
        "priority": 85,
    },
    "egyebszukseges1": {
        # PHOTOFALLBACK: ha van külön PNG/JPG fotó, akkor a kérelem PDF ide kerül.
        "canonical": "egyéb 1 - kérelem PDF, ha a fotó sorba tényleges fotó kerül",
        "aliases": ["kerelem", "kérelem", "tarteng", "tart eng", "tart.eng", "tart engedely", "tart engedély", "tartozkodasi engedely", "tartózkodási engedély", "kerelmi", "kérelmi"],
        "target_name": "kérelem",
        "priority": 90,
    },
    "egyebszukseges4": {
        "canonical": str(FIXED_DOCS["egyebszukseges4"]["canonical"]),
        "aliases": [],
        "target_name": str(FIXED_DOCS["egyebszukseges4"]["target_name"]),
        "priority": int(FIXED_DOCS["egyebszukseges4"]["priority"]),
    },
    "egyebszukseges5": {
        "canonical": str(FIXED_DOCS["egyebszukseges5"]["canonical"]),
        "aliases": [],
        "target_name": str(FIXED_DOCS["egyebszukseges5"]["target_name"]),
        "priority": int(FIXED_DOCS["egyebszukseges5"]["priority"]),
    },
}

CATEGORY_PRIORITY = {k: int(v["priority"]) for k, v in DOC_RULES.items()}

@dataclass
class ClassifiedFile:
    path: str
    filename: str
    doktip: str
    canonical: str
    score: int
    reason: str
    ext: str


def strip_accents(s: str) -> str:
    return "".join(ch for ch in unicodedata.normalize("NFKD", s) if not unicodedata.combining(ch))


def norm(s: str) -> str:
    s = strip_accents(str(s)).lower()
    s = re.sub(r"[_\-.,;:()\[\]{}]+", " ", s)
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def similarity(a: str, b: str) -> int:
    return int(round(100 * SequenceMatcher(None, norm(a), norm(b)).ratio()))


def abs_url(href: str) -> str:
    if href.startswith("http"):
        return href
    if href.startswith("/"):
        return BASE_URL + href
    return BASE_URL + "/eh/" + href.lstrip("./")


def classify_file(path: Path) -> Optional[ClassifiedFile]:
    ext = path.suffix.lower()
    if ext not in SUPPORTED_EXT:
        return None

    n = norm(path.stem)

    # v9: PNG/JPG fotó csak akkor javasolt, ha tényleg fotó/arckép jellegű fájlnév.
    # Ha van ilyen kép, a policy később ezt teszi az arckep sorba, a kérelem PDF-et pedig Egyéb 1-be.
    if ext in {".png", ".jpg", ".jpeg"}:
        photo_words = ["foto", "photo", "arckep", "kep", "profil", "portrait"]
        score = 100 if any(w in n for w in photo_words) else 80
        return ClassifiedFile(
            str(path),
            path.name,
            "arckep",
            "arckép/fotó kép",
            score,
            "kép fájl fotóként kezelve",
            ext,
        )

    best: Optional[Tuple[int, str, str]] = None

    for doktip, rule in DOC_RULES.items():
        aliases = list(rule["aliases"])  # type: ignore[index]
        for alias in aliases:
            an = norm(alias)
            score = 0
            reason = ""
            if an and an in n:
                score = 100
                reason = f"alias: {alias}"
            else:
                parts = n.split()
                windows = []
                for i in range(len(parts)):
                    for w in (1, 2, 3, 4):
                        windows.append(" ".join(parts[i:i+w]))
                fuzzy = max([similarity(an, c) for c in windows] + [similarity(an, n)]) if an else 0
                if fuzzy >= 78:
                    score = fuzzy
                    reason = f"fuzzy: {alias}"

            # Kérelem ne vigye el az előzetes megállapodást: ha előzetes van benne, az erősebb.
            if doktip == "egyebszukseges1" and ("elozetes" in n or "elomegallapodas" in n or "megallapodas" in n):
                score = min(score, 50)

            if score:
                rank = (score, -CATEGORY_PRIORITY.get(doktip, 999))
                if best is None or rank > (best[0], -CATEGORY_PRIORITY.get(best[1], 999)):
                    best = (score, doktip, reason)

    if not best or best[0] < 55:
        return None
    score, doktip, reason = best
    return ClassifiedFile(str(path), path.name, doktip, str(DOC_RULES[doktip]["canonical"]), score, reason, ext)


def scan_folder(folder: Path) -> List[ClassifiedFile]:
    files = [p for p in folder.rglob("*") if p.is_file() and p.suffix.lower() in SUPPORTED_EXT]
    out: List[ClassifiedFile] = []
    for f in files:
        c = classify_file(f)
        if c:
            out.append(c)
        else:
            out.append(ClassifiedFile(str(f), f.name, "__UNKNOWN__", "NEM FELISMERT", 0, "nincs erős fájlnév-egyezés", f.suffix.lower()))
    return out


def choose_best_per_doktip(items: List[ClassifiedFile]) -> Dict[str, ClassifiedFile]:
    chosen: Dict[str, ClassifiedFile] = {}
    for item in items:
        if item.doktip == "__UNKNOWN__" or item.score < MIN_SCORE_AUTO:
            continue
        old = chosen.get(item.doktip)
        if old is None or (item.score, -CATEGORY_PRIORITY.get(item.doktip, 999)) > (old.score, -CATEGORY_PRIORITY.get(old.doktip, 999)):
            chosen[item.doktip] = item
    return chosen


def is_image_file(item: ClassifiedFile) -> bool:
    return item.ext.lower() in {".png", ".jpg", ".jpeg"}


def is_pdf_file(item: ClassifiedFile) -> bool:
    return item.ext.lower() == ".pdf"


def looks_like_kerelem_pdf(item: ClassifiedFile) -> bool:
    if not is_pdf_file(item):
        return False
    n = norm(Path(item.filename).stem)
    if any(x in n for x in ["elozetes", "elomegallapodas", "megallapodas", "meghatalmaz", "utlevel", "szallas", "szallash", "posta", "oif"]):
        return False
    return any(x in n for x in ["kerelem", "kerelmi", "tarteng", "tart eng", "tart engedely", "tartozkodasi engedely"])


def apply_photo_kerelem_policy(chosen: Dict[str, ClassifiedFile], items: List[ClassifiedFile]) -> None:
    """
    Üzleti szabály v9:
    - Ha van külön PNG/JPG fotó a személymappában, akkor az arckep/fotó sorba az kerül,
      a kérelem PDF pedig egyebszukseges1-be.
    - Ha nincs külön PNG/JPG fotó, akkor marad a korábbi szabály: a kérelem PDF megy az arckep/fotó sorba,
      és egyebszukseges1 nem kapja meg ugyanazt duplikálva.
    """
    image_candidates = [i for i in items if i.doktip == "arckep" and is_image_file(i) and i.score >= MIN_SCORE_AUTO]
    image_candidates.sort(key=lambda i: (i.score, -len(i.filename)), reverse=True)

    kerelem_candidates = [i for i in items if looks_like_kerelem_pdf(i) and i.score >= 55]
    # Erősebb, ha a classifier is arckep/egyeb1 kérelemként ismerte fel.
    kerelem_candidates.sort(key=lambda i: ((1 if i.doktip in {"arckep", "egyebszukseges1"} else 0), i.score, -len(i.filename)), reverse=True)

    if image_candidates:
        photo = image_candidates[0]
        chosen["arckep"] = photo
        print(f"[POLICY] Van fotó kép: arckep <= {photo.filename}")
        if kerelem_candidates:
            kerelem = kerelem_candidates[0]
            chosen["egyebszukseges1"] = ClassifiedFile(
                kerelem.path,
                kerelem.filename,
                "egyebszukseges1",
                str(DOC_RULES["egyebszukseges1"]["canonical"]),
                max(kerelem.score, 100),
                "v9 policy: fotó kép van, ezért a kérelem PDF Egyéb 1-be kerül",
                kerelem.ext,
            )
            print(f"[POLICY] Kérelem PDF Egyéb 1-be: egyebszukseges1 <= {kerelem.filename}")
        return

    if kerelem_candidates:
        kerelem = kerelem_candidates[0]
        chosen["arckep"] = ClassifiedFile(
            kerelem.path,
            kerelem.filename,
            "arckep",
            str(DOC_RULES["arckep"]["canonical"]),
            max(kerelem.score, 100),
            "v9 policy: nincs fotó kép, ezért kérelem PDF megy a fotó/arckép sorba",
            kerelem.ext,
        )
        # Ne duplikáljuk ugyanazt a kérelmet Egyéb 1-be, ha nincs külön fotó kép.
        if chosen.get("egyebszukseges1") and Path(chosen["egyebszukseges1"].path) == Path(kerelem.path):
            del chosen["egyebszukseges1"]
        print(f"[POLICY] Nincs fotó kép: arckep <= {kerelem.filename}")


def find_projects_root_from(path: Path) -> Optional[Path]:
    """Megkeresi a szülők között a Projektek (jelöltek teljes anyagai) mappát."""
    try:
        candidates = [path if path.is_dir() else path.parent] + list((path if path.is_dir() else path.parent).parents)
    except Exception:
        candidates = []
    for p in candidates:
        if "projektek" in norm(p.name) and "jeloltek" in norm(p.name) and "teljes" in norm(p.name):
            return p
    return None


def unique_existing_dirs(paths: List[Optional[Path]]) -> List[Path]:
    out: List[Path] = []
    seen = set()
    for p in paths:
        if not p:
            continue
        try:
            pp = p.resolve()
        except Exception:
            pp = p
        if pp in seen:
            continue
        if pp.exists() and pp.is_dir():
            out.append(pp)
            seen.add(pp)
    return out


def candidate_fixed_roots(person_folder: Path, project_root: Optional[str] = None) -> List[Path]:
    roots: List[Optional[Path]] = []
    if project_root:
        pr = Path(str(project_root).strip('"'))
        roots.append(pr)
        roots.append(find_projects_root_from(pr))
        if pr.exists():
            roots.extend(list(pr.parents)[:4])
    roots.append(person_folder)
    roots.append(find_projects_root_from(person_folder))
    roots.extend(list(person_folder.parents)[:6])
    roots.append(Path.cwd())
    roots.append(Path(__file__).resolve().parent)
    return unique_existing_dirs(roots)


def find_fixed_file_for_doktip(doktip: str, person_folder: Path, project_root: Optional[str] = None) -> Optional[Path]:
    rule = FIXED_DOCS.get(doktip)
    if not rule:
        return None
    wanted_names = [str(x) for x in rule.get("filenames", [])]  # type: ignore[union-attr]
    wanted_norms = {norm(Path(x).stem) for x in wanted_names}
    roots = candidate_fixed_roots(person_folder, project_root)

    # 1) pontos / normalizált fájlnév közvetlenül a gyökerekben
    for root in roots:
        for name in wanted_names:
            p = root / name
            if p.exists() and p.is_file():
                return p
        try:
            for p in root.iterdir():
                if p.is_file() and p.suffix.lower() == ".pdf" and norm(p.stem) in wanted_norms:
                    return p
        except Exception:
            pass

    # 2) korlátozott rekurzív keresés. Először csak 3 szintig, hogy a nagy Drive ne lassuljon be nagyon.
    for root in roots:
        root_depth = len(root.parts)
        try:
            for p in root.rglob("*.pdf"):
                if len(p.parts) - root_depth > 3:
                    continue
                if norm(p.stem) in wanted_norms:
                    return p
        except Exception:
            pass
    return None


def add_fixed_documents(chosen: Dict[str, ClassifiedFile], person_folder: Path, project_root: Optional[str] = None) -> None:
    for doktip, rule in FIXED_DOCS.items():
        fp = find_fixed_file_for_doktip(doktip, person_folder, project_root)
        if not fp:
            print(f"[WARN] Fix dokumentum nem található ehhez: {doktip} / {rule.get('target_name')}")
            print("       Keresett helyek: személymappa szülői, Projektek (jelöltek teljes anyagai), script mappa.")
            print("       Tipp: tedd a PDF-et a 'Projektek (jelöltek teljes anyagai)' mappába vagy az EH Program mappába.")
            continue
        chosen[doktip] = ClassifiedFile(
            str(fp),
            fp.name,
            doktip,
            str(rule.get("canonical", doktip)),
            100,
            "fix mindenkinek feltöltendő dokumentum",
            fp.suffix.lower(),
        )
        print(f"[FIX] {doktip}: {fp}")

def find_person_folder(root: Path, person_name: str) -> Optional[Path]:
    if root.is_file():
        return root.parent
    if any(p.is_file() and p.suffix.lower() in SUPPORTED_EXT for p in root.iterdir()):
        return root
    target = norm(person_name)
    candidates: List[Tuple[int, Path]] = []
    root_depth = len(root.parts)
    for p in root.rglob("*"):
        if not p.is_dir():
            continue
        if len(p.parts) - root_depth > 4:
            continue
        pn = norm(p.name)
        score = similarity(target, pn)
        for t in target.split():
            if len(t) > 2 and t in pn:
                score += 5
        candidates.append((score, p))
    candidates.sort(reverse=True, key=lambda x: x[0])
    if candidates and candidates[0][0] >= 65:
        return candidates[0][1]
    return None


def infer_person_name_from_folder(folder: Path) -> str:
    """Személynév becslése a mappanévből.

    Támogatott minták:
      ARGUILLA MICHELLE
      ARGUILLA MICHELLE_1234567890
      ARGUILLA MICHELLE_adoazonosito
      ARGUILLA_MICHELLE_1234567890
    """
    name = folder.stem if folder.is_file() else folder.name
    name = name.replace("-", " ").strip()
    # Ha alulvonallal van adóazonosító / technikai suffix, vágjuk le.
    parts = [x.strip() for x in name.split("_") if x.strip()]
    if len(parts) >= 2:
        last = norm(parts[-1])
        digit_count = sum(ch.isdigit() for ch in parts[-1])
        # utolsó rész valószínű adóazonosító / technikai azonosító
        if digit_count >= 6 or last in {"ado", "adoazonosito", "adoszam", "adoazonosito jel"}:
            name = " ".join(parts[:-1])
        else:
            name = " ".join(parts)
    name = re.sub(r"\s+", " ", name).strip()
    return name


def extract_zip(zip_path: Path) -> Path:
    tmp = Path(tempfile.mkdtemp(prefix="eh_melleklet_zip_"))
    with zipfile.ZipFile(zip_path) as z:
        z.extractall(tmp)
    # Ha a ZIP gyökérben vannak a fájlok, tmp a személymappa.
    return tmp


def log_csv(path: Path, row: Dict[str, object]) -> None:
    exists = path.exists()
    with path.open("a", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=list(row.keys()))
        if not exists:
            w.writeheader()
        w.writerow(row)


# ---- v4: Excel / mappa / név alapú automatikus EH case ID keresés ----

def find_sheet_by_norm(wb, wanted: str):
    wn = norm(wanted)
    for name in wb.sheetnames:
        if norm(name) == wn:
            return wb[name]
    return None


def read_excel_person_context(workbook_path: Path, row_no: Optional[int] = None) -> Dict[str, str]:
    """Start!C1 + 1 alapján kiolvassa a személy nevét és a típust a Munkavállalók fülről."""
    try:
        import openpyxl
    except Exception:
        print("HIÁNYZÓ FÜGGŐSÉG: openpyxl nincs telepítve. Futtasd: py -m pip install openpyxl")
        raise

    wb = openpyxl.load_workbook(workbook_path, data_only=True, keep_vba=True, read_only=True)
    ws = find_sheet_by_norm(wb, "Munkavállalók")
    if ws is None:
        raise RuntimeError("Nem találom a Munkavállalók fület az Excelben.")

    if row_no is None:
        st = find_sheet_by_norm(wb, "Start")
        if st is not None:
            try:
                row_no = int(st["C1"].value) + 1
            except Exception:
                row_no = 2
        else:
            row_no = 2

    headers: Dict[str, int] = {}
    originals: Dict[str, str] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v is not None and str(v).strip():
            headers[norm(str(v))] = c
            originals[norm(str(v))] = str(v).strip()

    def val_by_headers(names: List[str]) -> str:
        for n in names:
            key = norm(n)
            if key in headers:
                v = ws.cell(row_no, headers[key]).value
                if v is not None and str(v).strip():
                    return " ".join(str(v).split())
        # részleges keresés
        for key, col in headers.items():
            for n in names:
                nk = norm(n)
                if nk and (nk in key or key in nk):
                    v = ws.cell(row_no, col).value
                    if v is not None and str(v).strip():
                        return " ".join(str(v).split())
        return ""

    full_name = val_by_headers(["Teljes Név", "Teljes név", "Név", "Munkavállaló neve", "Munkavallalo neve"])
    if not full_name:
        family = val_by_headers(["Családi név", "Családi név (útlevél szerint)", "Vezetéknév", "Vezeteknev", "kerelmezocsaladnev"])
        given = val_by_headers(["Utónév", "Utónév (útlevél szerint)", "Keresztnév", "Keresztnev", "kerelmezoutonev"])
        full_name = " ".join(x for x in [family, given] if x).strip()
    if not full_name:
        full_name = f"Excel sor {row_no}"

    def cell_letter(col: str) -> str:
        try:
            v = ws[f"{col}{row_no}"].value
            return "" if v is None else str(v).strip()
        except Exception:
            return ""

    de = cell_letter("DE")
    df = cell_letter("DF")
    app_type = ""
    if df == "36":
        app_type = "nemzeti"
    elif de == "1":
        app_type = "vendeg"

    ado = val_by_headers(["Adóazonosító", "Adoazonosito", "Adóazonosító jel", "Adó szám", "Adószám"])

    return {
        "row_no": str(row_no),
        "person_name": full_name,
        "app_type": app_type,
        "adoazonosito": ado,
        "de": de,
        "df": df,
    }


def case_type_matches(row_text: str, app_type: str) -> bool:
    t = norm(row_text)
    if not app_type:
        return True
    if app_type == "nemzeti":
        return "nemzeti" in t or "kartya" in t
    if app_type == "vendeg":
        return "vendegmunkas" in t or "vendegmunkas" in strip_accents(row_text).lower() or "vendégmunkás" in row_text.lower()
    return True


def case_match_score(person_name: str, row_text: str, app_type: str = "") -> int:
    pn = norm(person_name)
    rt = norm(row_text)
    if not pn or pn.startswith("excel sor"):
        return 0
    tokens = [x for x in pn.split() if len(x) > 1]
    if pn in rt:
        score = 100
    else:
        score = similarity(pn, rt)
        hits = sum(1 for tok in tokens if tok in rt)
        if tokens and hits == len(tokens):
            score = max(score, 92)
        elif tokens and hits >= max(1, len(tokens) - 1):
            score = max(score, 80)
    if app_type and case_type_matches(row_text, app_type):
        score += 4
    return min(score, 100)


def safe_goto(page, url: str, *, timeout: int = 60000, attempts: int = 3) -> None:
    """
    EH néha login/session után ugyanarra az URL-re második navigációt indít,
    amit a Playwright "interrupted by another navigation" hibának lát.
    Ez itt nem valódi hiba: megvárjuk, hogy lecsengjen, és csak akkor dobunk hibát,
    ha több próbálkozás után sem jutunk el az oldalra.
    """
    last_error = None
    target_path = url.replace(BASE_URL, "")

    for attempt in range(1, attempts + 1):
        try:
            # Ha már jó oldalon vagyunk, ne indítsunk felesleges második navigációt.
            if target_path and target_path in page.url:
                try:
                    page.wait_for_load_state("domcontentloaded", timeout=8000)
                except Exception:
                    pass
                return

            page.goto(url, wait_until="domcontentloaded", timeout=timeout)
            return

        except Exception as exc:
            last_error = exc
            msg = str(exc)
            soft_navigation_error = (
                "interrupted by another navigation" in msg
                or "Navigation failed because page was closed" in msg
                or "net::ERR_ABORTED" in msg
                or "Execution context was destroyed" in msg
            )
            if not soft_navigation_error:
                raise

            print(f"[NAV-WARN] EH navigáció megszakadt, újrapróba {attempt}/{attempts}: {msg.splitlines()[0]}")
            time.sleep(1.0)
            try:
                page.wait_for_load_state("domcontentloaded", timeout=12000)
            except Exception:
                pass

            # Ha a megszakítás ellenére a kívánt EH oldalra kerültünk, jók vagyunk.
            if target_path and target_path in page.url:
                return

    if last_error:
        raise last_error


def is_login_page(page) -> bool:
    try:
        return "login" in page.url.lower() or page.locator("input[type='password']").count() > 0
    except Exception:
        return "login" in page.url.lower()


def wait_for_cases_login(page) -> None:
    url = f"{BASE_URL}/eh/cases"
    safe_goto(page, url)
    time.sleep(1)

    if is_login_page(page):
        print("\nBelépés szükséges. Jelentkezz be az EH-ban, majd itt nyomj Entert.")
        input("Enter, ha beléptél...")
        # Login után az EH gyakran magától navigál. Előbb várunk, aztán csak akkor megyünk /eh/cases-re, ha kell.
        try:
            page.wait_for_load_state("domcontentloaded", timeout=12000)
        except Exception:
            pass
        time.sleep(1)
        if "/eh/cases" not in page.url:
            safe_goto(page, url)
        else:
            # Már cases oldalon vagyunk, nem indítunk új page.goto-t ugyanarra, ez okozta a v4 hibát.
            try:
                page.wait_for_load_state("domcontentloaded", timeout=8000)
            except Exception:
                pass
        time.sleep(1)

    # Végső ellenőrzés: ha még mindig login oldalon vagyunk, adjunk egy egyértelmű üzenetet.
    if is_login_page(page):
        print("[WARN] Úgy tűnik, még mindig login oldalon vagy. Lépj be, majd futtasd újra vagy nyomj Entert, ha már beléptél.")
        input("Enter folytatáshoz...")
        safe_goto(page, url)
        time.sleep(1)


def extract_case_candidates_from_page(page, person_name: str, app_type: str = "") -> List[Dict[str, object]]:
    candidates: List[Dict[str, object]] = []
    rows = page.locator("tr[data-kid]")
    try:
        count = rows.count()
    except Exception:
        count = 0
    for i in range(count):
        row = rows.nth(i)
        try:
            kid = row.get_attribute("data-kid") or ""
            text = " ".join(row.inner_text(timeout=1500).split())
        except Exception:
            continue
        if not kid:
            m = re.search(r"EH(\d{6,})", text)
            kid = m.group(1) if m else ""
        if not kid:
            continue
        score = case_match_score(person_name, text, app_type)
        if score >= 65:
            candidates.append({"case_id": kid, "score": score, "text": text})
    candidates.sort(key=lambda x: int(x["score"]), reverse=True)
    return candidates


def find_case_id_on_eh(page, person_name: str, app_type: str = "") -> Optional[str]:
    wait_for_cases_login(page)
    seen_urls = set()
    all_candidates: List[Dict[str, object]] = []

    for _ in range(6):
        if page.url in seen_urls:
            break
        seen_urls.add(page.url)
        all_candidates.extend(extract_case_candidates_from_page(page, person_name, app_type))
        if all_candidates and int(all_candidates[0]["score"]) >= 95:
            break

        # Ha van következő oldal link, próbáljuk meg. Ha nincs, kilépünk.
        clicked = False
        for sel in ["a:has-text('Következő')", "a:has-text('következő')", "a[rel='next']", ".pagination a:has-text('>')"]:
            try:
                loc = page.locator(sel).first
                if loc.count():
                    loc.click(timeout=2000)
                    page.wait_for_load_state("domcontentloaded", timeout=10000)
                    time.sleep(0.5)
                    clicked = True
                    break
            except Exception:
                continue
        if not clicked:
            break

    # Duplikátumok kiszűrése
    unique: Dict[str, Dict[str, object]] = {}
    for c in all_candidates:
        cid = str(c["case_id"])
        if cid not in unique or int(c["score"]) > int(unique[cid]["score"]):
            unique[cid] = c
    candidates = sorted(unique.values(), key=lambda x: int(x["score"]), reverse=True)

    if not candidates:
        print(f"[WARN] Nem találtam EH ügyet erre a névre: {person_name}")
        return None

    print("\nEH ügyjelöltek:")
    for idx, c in enumerate(candidates[:8], start=1):
        print(f"  {idx}. EH{c['case_id']} | score={c['score']} | {c['text'][:180]}")

    if len(candidates) == 1 or int(candidates[0]["score"]) >= 95 and (len(candidates) == 1 or int(candidates[0]["score"]) - int(candidates[1]["score"]) >= 8):
        cid = str(candidates[0]["case_id"])
        print(f"[AUTO CASE] {person_name} -> EH{cid}")
        return cid

    ans = input("Melyik legyen? Írd be a sorszámot, konkrét case ID-t, vagy Enter = 1.: ").strip()
    if not ans:
        return str(candidates[0]["case_id"])
    if ans.isdigit():
        n = int(ans)
        if 1 <= n <= len(candidates[:8]):
            return str(candidates[n - 1]["case_id"])
        if len(ans) >= 6:
            return ans
    return None


def click_mellekletek_tab(page) -> None:
    try:
        page.evaluate("""() => {
          const a = document.querySelector('a[href$="#mellekletek"], a[href*="#mellekletek"]');
          if (a) a.click();
        }""")
    except Exception:
        pass
    time.sleep(0.5)


def wait_for_logged_in(page, case_id: str) -> None:
    wanted = f"{BASE_URL}/eh/cases/edit/{case_id}#mellekletek"
    page.goto(wanted, wait_until="domcontentloaded", timeout=60000)
    time.sleep(1)
    if "login" in page.url.lower() or page.locator("input[type='password']").count() > 0:
        print("\nBelépés szükséges. Jelentkezz be a megnyílt böngészőben.")
        input("Ha bent vagy az EH-ban, nyomj Entert itt...")
        page.goto(wanted, wait_until="domcontentloaded", timeout=60000)
    click_mellekletek_tab(page)


def parse_attachment_rows(page) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    locs = page.locator("#mellekletek tr[doktip], tr[doktip]")
    for i in range(locs.count()):
        row = locs.nth(i)
        doktip = row.get_attribute("doktip") or ""
        try:
            label = " ".join(row.locator("td").first.inner_text(timeout=1500).split())
        except Exception:
            label = ""
        select_name = ""
        try:
            if row.locator("select").count():
                select_name = row.locator("select").first.get_attribute("name") or ""
        except Exception:
            pass
        upload_href = ""
        try:
            a = row.locator("a:has-text('Feltölt')").first
            if a.count():
                upload_href = a.get_attribute("href") or ""
        except Exception:
            pass
        rows.append({"doktip": doktip, "label": label, "select_name": select_name, "upload_href": upload_href})
    # Egyedi doktip sorrendben.
    seen = set()
    uniq = []
    for r in rows:
        if r["doktip"] in seen:
            continue
        seen.add(r["doktip"])
        uniq.append(r)
    return uniq


def option_texts_for_doktip(page, doktip: str) -> List[Tuple[str, str]]:
    script = """
    (doktip) => {
      const row = document.querySelector(`#mellekletek tr[doktip="${doktip}"]`) || document.querySelector(`tr[doktip="${doktip}"]`);
      if (!row) return [];
      const sel = row.querySelector('select');
      if (!sel) return [];
      return Array.from(sel.options).map(o => [o.value, o.textContent.trim()]);
    }
    """
    try:
        return [(str(v), str(t)) for v, t in page.evaluate(script, doktip)]
    except Exception:
        return []


def select_existing_option(page, doktip: str, file_path: Path) -> bool:
    stem_n = norm(file_path.stem)
    file_n = norm(file_path.name)
    opts = option_texts_for_doktip(page, doktip)
    best = ("", "", 0)
    for value, text in opts:
        if value in {"0", ""}:
            continue
        tn = norm(text)
        score = max(similarity(stem_n, tn), similarity(file_n, tn))
        # Erős plusz, ha a stem több tokenje benne van az EH option textben.
        tokens = [t for t in stem_n.split() if len(t) > 2]
        hits = sum(1 for t in tokens if t in tn)
        if tokens and hits >= max(1, min(3, len(tokens))):
            score = max(score, 88 + min(10, hits))
        if stem_n and stem_n in tn:
            score = 100
        if score > best[2]:
            best = (value, text, score)
    if best[0] and best[2] >= 70:
        page.evaluate(
            """([doktip, value]) => {
                const row = document.querySelector(`#mellekletek tr[doktip="${doktip}"]`) || document.querySelector(`tr[doktip="${doktip}"]`);
                const sel = row && row.querySelector('select');
                if (sel) {
                    sel.value = value;
                    sel.dispatchEvent(new Event('change', {bubbles:true}));
                }
            }""",
            [doktip, best[0]],
        )
        print(f"[SELECT] {doktip}: {best[1]} (score {best[2]})")
        return True
    return False


def click_first_available(page, selectors: List[str], timeout: int = 5000) -> bool:
    for sel in selectors:
        try:
            loc = page.locator(sel).first
            if loc.count():
                loc.click(timeout=timeout)
                return True
        except Exception:
            continue
    return False


def upload_file_for_doktip(page, case_id: str, doktip: str, file_path: Path) -> bool:
    page.goto(f"{BASE_URL}/eh/cases/edit/{case_id}#mellekletek", wait_until="domcontentloaded", timeout=60000)
    click_mellekletek_tab(page)
    rows = parse_attachment_rows(page)
    row = next((r for r in rows if r["doktip"] == doktip), None)
    if not row:
        print(f"[SKIP] EH oldalon nincs ilyen doktip: {doktip} ({file_path.name})")
        return False

    if select_existing_option(page, doktip, file_path):
        return True

    href = row.get("upload_href") or f"/eh/storage?add={doktip}&back=cases/edit/{case_id}&melleklet"
    url = abs_url(href)
    print(f"[UPLOAD] {doktip}: {file_path.name}")
    page.goto(url, wait_until="domcontentloaded", timeout=60000)
    time.sleep(0.7)

    file_input = page.locator("input[type='file']").first
    try:
        file_input.wait_for(state="attached", timeout=15000)
        file_input.set_input_files(str(file_path))
    except PlaywrightTimeoutError:
        print(f"[MANUAL] Nem találtam file inputot a storage oldalon: {page.url}")
        input("Töltsd fel kézzel ezt a fájlt, majd Enter: " + file_path.name)
        page.goto(f"{BASE_URL}/eh/cases/edit/{case_id}#mellekletek", wait_until="domcontentloaded", timeout=60000)
        click_mellekletek_tab(page)
        return select_existing_option(page, doktip, file_path)

    # Megnevezés/név mező, ha létezik.
    for sel in [
        "input[name*='megnevezes']",
        "input[name*='megnevezés']",
        "input[name='nev']",
        "input[name*='name']",
        "input[type='text']",
    ]:
        try:
            loc = page.locator(sel)
            if loc.count():
                first = loc.first
                val = first.input_value(timeout=800)
                if not val:
                    first.fill(file_path.stem[:120])
                break
        except Exception:
            pass

    clicked = click_first_available(page, [
        "button[type='submit']",
        "input[type='submit']",
        "button:has-text('Feltölt')",
        "input[value*='Feltölt']",
        "button:has-text('Mentés')",
        "button:has-text('Mehet')",
        "a.btn:has-text('Feltölt')",
        "a.btn:has-text('Mentés')",
    ], timeout=10000)
    if not clicked:
        print("[MANUAL] Nem találtam feltöltő submit gombot.")
        input("Kattints kézzel a feltöltésre, majd Enter...")

    try:
        page.wait_for_load_state("domcontentloaded", timeout=60000)
    except Exception:
        pass
    time.sleep(1.5)
    page.goto(f"{BASE_URL}/eh/cases/edit/{case_id}#mellekletek", wait_until="domcontentloaded", timeout=60000)
    click_mellekletek_tab(page)
    time.sleep(1)
    if select_existing_option(page, doktip, file_path):
        return True
    print(f"[WARN] Feltöltés után sem találtam kiválasztható opciót: {doktip} / {file_path.name}")
    input("Ha kézzel látod a selectben, válaszd ki, majd Enter...")
    return True


def submit_attachment_form(page) -> None:
    page.evaluate("""() => {
        const form = document.querySelector('form.melleklet');
        if (form) form.submit();
    }""")
    try:
        page.wait_for_load_state("domcontentloaded", timeout=30000)
    except Exception:
        pass


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--case-id", help="EH ügy ID, pl. 19020197. Ha nincs megadva, név alapján keresi az EH Ügyeim listában.")
    ap.add_argument("--person-folder", help="Konkrét személymappa lokális útvonala. Ha nincs case-id, a mappanévből kinyeri a nevet és EH Ügyeim alapján keres case ID-t.")
    ap.add_argument("--project-root", "--root", dest="root", help="Projekt / Google Drive lokális root")
    ap.add_argument("--fixed-root", help="Fix dokumentumok mappája. Ha nincs megadva, a script a Projektek (jelöltek teljes anyagai) / szülő mappák / script mappa között keresi.")
    ap.add_argument("--person-name", help="Személy neve root és EH case kereséshez")
    ap.add_argument("--excel", "--workbook", dest="excel", help="XLSM elérési út; Start!C1 alapján olvas nevet és típust")
    ap.add_argument("--row", type=int, help="Excel fizikai sor száma a Munkavállalók fülön. Ha nincs: Start!C1 + 1.")
    ap.add_argument("--type", choices=["nemzeti", "vendeg"], help="Kérelemtípus felülírása az EH case találatok szűréséhez")
    ap.add_argument("--zip", dest="zip_path", help="ZIP fájl, ha Drive-ból letöltött csomaggal tesztelsz")
    ap.add_argument("--profile", default=".eh_edge_profile", help="Playwright böngészőprofil mappa")
    ap.add_argument("--no-submit", action="store_true", help="Ne küldje be a melléklet formot")
    ap.add_argument("--dry-run", action="store_true", help="Csak felismerés, EH nélkül")
    args = ap.parse_args()

    excel_ctx: Dict[str, str] = {}
    if args.excel:
        excel_ctx = read_excel_person_context(Path(args.excel), args.row)
        if not args.person_name:
            args.person_name = excel_ctx.get("person_name") or ""
        if not args.type and excel_ctx.get("app_type"):
            args.type = excel_ctx.get("app_type")  # type: ignore[assignment]
        print("\nEXCEL KONTEXTUS")
        print(f"  Excel: {args.excel}")
        print(f"  Sor: {excel_ctx.get('row_no')}")
        print(f"  Név: {args.person_name}")
        print(f"  Típus: {args.type or 'nem ismert'}")

    temp_dir: Optional[Path] = None
    if args.zip_path:
        temp_dir = extract_zip(Path(args.zip_path))
        person_folder = temp_dir
    elif args.person_folder:
        person_folder = Path(args.person_folder.strip('"'))
        if not args.person_name:
            args.person_name = infer_person_name_from_folder(person_folder)
            print(f"[INFO] Személynév a mappanévből: {args.person_name}")
    else:
        root_s = args.root or input("Lokális Drive/projekt root vagy személymappa útvonala: ").strip().strip('"')
        root = Path(root_s)
        if root.suffix.lower() == ".zip":
            temp_dir = extract_zip(root)
            person_folder = temp_dir
        else:
            person_name = args.person_name or input("Munkavállaló neve a mappa kereséséhez: ").strip()
            found = find_person_folder(root, person_name)
            if not found:
                print("Nem találtam személymappát.")
                return 3
            person_folder = found

    if not args.person_name and 'person_folder' in locals():
        args.person_name = infer_person_name_from_folder(person_folder)
        print(f"[INFO] Személynév a mappanévből: {args.person_name}")

    if not person_folder.exists():
        print(f"Nem létező mappa: {person_folder}")
        return 3

    print(f"\nSzemélymappa/forrás: {person_folder}")
    classified = scan_folder(person_folder)
    for c in classified:
        print(f"  {c.filename} -> {c.doktip} ({c.score}) {c.canonical} | {c.reason}")
    chosen = choose_best_per_doktip(classified)
    apply_photo_kerelem_policy(chosen, classified)
    add_fixed_documents(chosen, person_folder, args.fixed_root or args.root)
    if not chosen:
        print("Nincs automatikusan feltölthető dokumentum.")
        return 4

    print("\nAutomatikusan feltöltésre javasolt:")
    for doktip, c in sorted(chosen.items(), key=lambda kv: CATEGORY_PRIORITY.get(kv[0], 999)):
        print(f"  {doktip:34s} <= {Path(c.path).name}")

    report_path = Path.cwd() / "eh_melleklet_asszisztens_log.csv"

    if args.dry_run:
        print(f"\nDry-run kész. Log: {report_path}")
        return 0

    print("\nIndítás: EH megnyitása. Ha case ID nincs megadva, először megkeresem az Ügyeim listában.")
    profile_dir = Path.cwd() / args.profile
    with sync_playwright() as p:
        try:
            context = p.chromium.launch_persistent_context(
                user_data_dir=str(profile_dir),
                channel="msedge",
                headless=False,
                accept_downloads=True,
                ignore_https_errors=True,
                chromium_sandbox=True,
            )
        except Exception:
            context = p.chromium.launch_persistent_context(
                user_data_dir=str(profile_dir),
                headless=False,
                accept_downloads=True,
                ignore_https_errors=True,
                chromium_sandbox=True,
            )
        page = context.pages[0] if context.pages else context.new_page()
        page.set_default_timeout(8000)
        try:
            case_id = args.case_id
            if not case_id:
                person_name = args.person_name or input("Munkavállaló neve EH case kereséshez: ").strip()
                case_id = find_case_id_on_eh(page, person_name, args.type or "")
                if not case_id:
                    case_id = input("Nem sikerült automatikusan. Írd be kézzel az EH case ID-t: ").strip()
                if not re.fullmatch(r"\d{6,}", case_id or ""):
                    print("Hibás / hiányzó case ID.")
                    return 2

            print(f"\nEH case ID: {case_id}")
            for doktip, c in chosen.items():
                log_csv(report_path, {
                    "time": datetime.now().isoformat(timespec="seconds"),
                    "case_id": case_id,
                    "person_folder": str(person_folder),
                    "doktip": doktip,
                    "file": c.path,
                    "score": c.score,
                    "action": "planned",
                })

            wait_for_logged_in(page, str(case_id))
            rows = parse_attachment_rows(page)
            available = {r["doktip"] for r in rows}
            print("\nEH oldalon elérhető doktip-ek:")
            print(", ".join(sorted(available)))

            ok_count = 0
            target_items = [(d, c) for d, c in sorted(chosen.items(), key=lambda kv: CATEGORY_PRIORITY.get(kv[0], 999)) if d in available]
            skipped = [d for d in chosen if d not in available]
            for d in skipped:
                print(f"[SKIP] A live EH oldalon nincs ilyen melléklet sor: {d}")

            for doktip, c in target_items:
                fp = Path(c.path)
                ok = upload_file_for_doktip(page, str(case_id), doktip, fp)
                log_csv(report_path, {
                    "time": datetime.now().isoformat(timespec="seconds"),
                    "case_id": case_id,
                    "person_folder": str(person_folder),
                    "doktip": doktip,
                    "file": str(fp),
                    "score": c.score,
                    "action": "uploaded_or_selected" if ok else "failed",
                })
                if ok:
                    ok_count += 1

            page.goto(f"{BASE_URL}/eh/cases/edit/{case_id}#mellekletek", wait_until="domcontentloaded", timeout=60000)
            click_mellekletek_tab(page)
            print(f"\nFeltöltés/választás kész: {ok_count}/{len(target_items)} dokumentum.")
            if args.no_submit:
                print("NO-SUBMIT: nem rögzítem a melléklet formot.")
                input("Ellenőrzés után Enter a böngésző bezárásához...")
            else:
                ans = input("Ellenőrizd az EH Fájlmellékletek fület. Rögzítsem a mellékletválasztásokat? (igen/N): ").strip().lower()
                if ans in {"i", "igen", "y", "yes"}:
                    submit_attachment_form(page)
                    print("Melléklet form rögzítés elküldve.")
                else:
                    print("Nem rögzítettem. Kézzel még tudod ellenőrizni / rögzíteni.")
                input("Enter a böngésző bezárásához...")
        finally:
            try:
                context.close()
            except Exception:
                pass
    if temp_dir:
        try:
            shutil.rmtree(temp_dir, ignore_errors=True)
        except Exception:
            pass
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
